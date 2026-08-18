"""
Compare ETools-generated Casing Review Excel against existing (ground-truth) files.
Outputs a report to comparison_report.txt
"""
import sys, os, pathlib, io
sys.path.insert(0, str(pathlib.Path(__file__).parent / "app"))

import openpyxl
from openpyxl import load_workbook

out = io.open("comparison_report.txt", "w", encoding="utf-8")

def log(msg=""):
    print(msg)
    out.write(msg + "\n")
    out.flush()

APD_DIR = pathlib.Path(r"C:\ETools_Portable\APD")
OUTPUT_DIR = pathlib.Path(r"C:\ETools_Portable\app\output")

# The 3 matched pairs: (apd_pdf_path, existing_excel_path)
PAIRS = [
    (
        APD_DIR / "application_43013537270000.pdf",
        APD_DIR / "Casing Review_43013537270000_Myton City UT 16-23 3-2-25-36-7H.xlsx",
    ),
    (
        APD_DIR / "application_43013537010000 Check.pdf",
        APD_DIR / "Casing Review_43013537010000_UT 16-9 3-2-16-21-1H.xlsx",
    ),
    (
        APD_DIR / "application_13067.pdf",
        APD_DIR / "Casing Review_43019500930000_Federal 1-15H-20-21.xlsx",
    ),
]


def generate_excel(apd_path):
    """Run CasingReviewService.generate() and return path to generated Excel."""
    log(f"  Generating from {apd_path.name} ...")
    from etools.services.casing_review_service import CasingReviewService
    svc = CasingReviewService()
    result = svc.generate(apd_pdf_path=str(apd_path))
    log(f"  Generated: {result.output_path}")
    return pathlib.Path(result.output_path)


def load_sheet_values(wb, sheet_name):
    """Return {(row, col): value} for all non-None cells in a sheet."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    data = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                data[(cell.row, cell.column)] = cell.value
    return data


def compare_sheets(wb_gen, wb_exist, sheet_name):
    """Compare a single sheet between generated and existing workbooks. Return list of diffs."""
    gen_data = load_sheet_values(wb_gen, sheet_name)
    exist_data = load_sheet_values(wb_exist, sheet_name)

    diffs = []
    all_keys = set(gen_data.keys()) | set(exist_data.keys())
    for (r, c) in sorted(all_keys):
        gval = gen_data.get((r, c))
        eval_ = exist_data.get((r, c))
        if gval == eval_:
            continue
        # Both numeric: check if different beyond tolerance
        try:
            gf = float(gval) if gval is not None else None
            ef = float(eval_) if eval_ is not None else None
            if gf is not None and ef is not None:
                if abs(gf - ef) < 1e-6:
                    continue  # float equality within tolerance
                pct = abs(gf - ef) / max(abs(ef), 1e-9) * 100
                diffs.append({
                    "row": r, "col": c,
                    "generated": gval, "existing": eval_,
                    "type": "numeric", "pct_diff": pct
                })
                continue
        except (TypeError, ValueError):
            pass
        # String comparison
        gs = str(gval).strip() if gval is not None else ""
        es = str(eval_).strip() if eval_ is not None else ""
        if gs != es:
            diffs.append({
                "row": r, "col": c,
                "generated": gval, "existing": eval_,
                "type": "string", "pct_diff": None
            })
    return diffs


def colnum_to_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def report_pair(apd_path, exist_path):
    log("=" * 70)
    log(f"APD: {apd_path.name}")
    log(f"Existing Excel: {exist_path.name}")
    log("=" * 70)

    if not exist_path.exists():
        log("  ERROR: existing Excel not found — skipping")
        return

    gen_path = generate_excel(apd_path)
    if not gen_path.exists():
        log("  ERROR: generated Excel not found — generation may have failed")
        return

    log(f"  Loading workbooks ...")
    wb_gen = load_workbook(str(gen_path), data_only=True)
    wb_exist = load_workbook(str(exist_path), data_only=True)

    gen_sheets = set(wb_gen.sheetnames)
    exist_sheets = set(wb_exist.sheetnames)

    only_in_gen = gen_sheets - exist_sheets
    only_in_exist = exist_sheets - gen_sheets
    common = gen_sheets & exist_sheets

    if only_in_gen:
        log(f"  Sheets only in GENERATED: {sorted(only_in_gen)}")
    if only_in_exist:
        log(f"  Sheets only in EXISTING: {sorted(only_in_exist)}")

    total_diffs = 0
    KEY_SHEETS = ["Casing Review", "BOPE", "Formations", "DataPrint"]

    for sheet in sorted(common):
        diffs = compare_sheets(wb_gen, wb_exist, sheet)
        if not diffs:
            log(f"  [{sheet}] -- MATCH (no differences)")
            continue

        is_key = sheet in KEY_SHEETS
        log(f"\n  [{sheet}] -- {len(diffs)} difference(s){' ***KEY SHEET***' if is_key else ''}")
        shown = 0
        for d in diffs:
            cell = f"{colnum_to_letter(d['col'])}{d['row']}"
            if d["type"] == "numeric":
                pct_str = f" ({d['pct_diff']:.2f}% diff)"
                log(f"    {cell}: GEN={d['generated']!r}  EXIST={d['existing']!r}{pct_str}")
            else:
                g_short = repr(str(d['generated'])[:80]) if d['generated'] is not None else "None"
                e_short = repr(str(d['existing'])[:80]) if d['existing'] is not None else "None"
                log(f"    {cell}: GEN={g_short}  EXIST={e_short}")
            shown += 1
            if shown >= 40 and not is_key:
                log(f"    ... ({len(diffs) - shown} more diffs not shown for non-key sheet)")
                break
        total_diffs += len(diffs)

    log(f"\n  TOTAL DIFFERENCES: {total_diffs}")
    log("")


log("ETools Casing Review Excel Comparison Report")
log("=" * 70)
log("")

for apd_path, exist_path in PAIRS:
    try:
        report_pair(apd_path, exist_path)
    except Exception as e:
        import traceback
        log(f"ERROR processing {apd_path.name}: {e}")
        log(traceback.format_exc())
    log("")

out.close()
print("Report written to comparison_report.txt")
