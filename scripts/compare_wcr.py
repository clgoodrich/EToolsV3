"""Compare generated WCR workbooks against the hand-made originals.

Two wells in the repo have a hand-made WCR Excel to compare against:

  * South Moon 5-31-32-C4-3H (4301353996) — reference workbook in
    tests/fixtures/reference, WCR PDF in tests/fixtures/wcr. Regenerated
    through the primary WCRPdfService pipeline (PDF + DB survey).
  * Reay 16-29-30-B4-2H (4301354722) — hand-made workbook in archive/misc.
    No WCR PDF exists on disk, so it runs through the legacy From-Database
    flow (WellService -> SurveyService -> ClearanceService -> WCRService).

The diff aligns rows by label (info-block labels, location-row names)
rather than by position, because the hand-made workbooks shift rows
around (extra lateral row, header on row 10 vs 9, missing BHL row).

Usage:  python scripts/compare_wcr.py [south_moon] [reay]
        (no args = both)
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

OUT_DIR = ROOT / "output" / "wcr_compare"

SOUTH_MOON_API = "4301353996"
SOUTH_MOON_PDF = ROOT / "tests/fixtures/wcr/WCR 43013539960000.pdf"
SOUTH_MOON_REF = ROOT / "tests/fixtures/reference/South_Moon_5-31-32-C4-3H_4301353996_WCR.xlsx"

REAY_API = "4301354722"
REAY_REF = ROOT / "archive/misc/Reay_16-29-30-B4-2H_4301354722_WCR.xlsx"

INFO_LABELS = (
    "WellName", "API", "Operator", "Lateral", "ConstructKeyWellType",
    "WellType", "SpudDate", "RotaryRigDate", "TDReachedDate",
    "CompletedOrAbandonedDate",
)
LOCATION_NAMES = ("SHL", "Control_Point", "Frac_Start", "Frac_End", "BHL", "KOP", "Landing")
LOCATION_COLS = (
    "MD", "TVD", "Easting", "Northing", "FNL", "FSL", "FEL", "FWL",
    "Sec", "Twp", "TwpDir", "Rng", "RngDir", "Base",
)


# ---------------------------------------------------------------------------
# generation
# ---------------------------------------------------------------------------


def generate_south_moon() -> Path:
    from etools.repositories import SurveyRepository
    from etools.services.wcr_pdf_service import WCRPdfService

    surveys = SurveyRepository().get_points_by_api_lateral(SOUTH_MOON_API, "0000")["AsDrilled"]
    out = OUT_DIR / "South_Moon_4301353996_generated.xlsx"
    result = WCRPdfService().generate(
        wcr_pdf_path=SOUTH_MOON_PDF, surveys=surveys, output_path=out
    )
    return result.output_path


def generate_reay() -> Path:
    from etools.models import SurveyFrame, WellLookup
    from etools.services import WCRService
    from etools.services.clearance_service import ClearanceService
    from etools.services.survey_service import SurveyService
    from etools.services.well_service import WellService

    bundle = WellService().load(WellLookup(api=REAY_API, lateral="0000"))
    processed = SurveyService().process(bundle.headers, bundle.surveys)
    citing = "AsDrilled" if "AsDrilled" in processed else next(iter(processed))
    sr = processed[citing]
    clearance = ClearanceService().calculate(
        sr.frames[SurveyFrame.TRUE], kop_md=sr.kop.md, landing_md=sr.landing_md
    )
    return WCRService().generate(
        api=REAY_API,
        lateral="0000",
        summary_footages=clearance.summary,
        points=clearance.points,
        output_dir=OUT_DIR,
    )


# ---------------------------------------------------------------------------
# workbook -> labelled blocks
# ---------------------------------------------------------------------------


def parse_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    info: dict[str, object] = {}
    locations: dict[str, list] = {}
    perf: dict[str, object] = {}
    casing: list[list] = []

    casing_header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=15):
        label = row[0].value
        vals = [c.value for c in row]
        if label in INFO_LABELS:
            info[str(label)] = vals[1]
        elif label in LOCATION_NAMES:
            locations[str(label)] = vals[1:15]
        elif label == "Feature":
            casing_header_row = row[0].row
        elif casing_header_row and row[0].row > casing_header_row and any(
            v is not None for v in vals
        ):
            casing.append(vals[:13])
        # perf block lives in E1:G2 on hand-made sheets
        if row[0].row == 1 and vals[4] == "Perf Top":
            perf["header"] = True
        if row[0].row == 2 and perf.get("header"):
            perf["top"], perf["bottom"], perf["date"] = vals[4], vals[5], vals[6]
    return {"info": info, "locations": locations, "perf": perf, "casing": casing}


def _fmt(v) -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        return f"{v:,.2f}".rstrip("0").rstrip(".")
    if isinstance(v, int):
        return f"{v:,}"
    return str(v)


def _num_close(a, b, tol: float) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def diff(ref_path: Path, gen_path: Path, label: str) -> None:
    ref = parse_workbook(ref_path)
    gen = parse_workbook(gen_path)
    print(f"\n{'=' * 84}\n{label}")
    print(f"  hand-made: {ref_path.relative_to(ROOT)}")
    print(f"  generated: {gen_path.relative_to(ROOT)}\n{'=' * 84}")

    print("\n-- Info block (aligned by label) --")
    print(f"  {'label':<26} {'hand-made':<28} {'generated':<28}")
    for k in INFO_LABELS:
        rv, gv = ref["info"].get(k), gen["info"].get(k)
        if rv is None and gv is None:
            continue
        mark = "  " if str(rv) == str(gv) else "✗ "
        print(f"  {mark}{k:<24} {_fmt(rv):<28} {_fmt(gv):<28}")

    print("\n-- Location rows (aligned by name) --")
    for name in LOCATION_NAMES:
        rrow, grow = ref["locations"].get(name), gen["locations"].get(name)
        if rrow is None and grow is None:
            continue
        if rrow is None or grow is None:
            side = "hand-made" if grow is None else "generated"
            print(f"  {name}: only in {'hand-made' if grow is None else 'generated'} "
                  f"({side} row: { [_fmt(v) for v in (rrow or grow)[:8]] })")
            continue
        diffs = []
        for i, col in enumerate(LOCATION_COLS):
            rv, gv = rrow[i], grow[i]
            if rv is None and gv is None:
                continue
            if str(rv) == str(gv) or _num_close(rv, gv, 0.5):
                continue
            note = ""
            try:
                note = f" (Δ {abs(float(rv) - float(gv)):,.1f})"
            except (TypeError, ValueError):
                pass
            diffs.append(f"{col}: {_fmt(rv)} vs {_fmt(gv)}{note}")
        if diffs:
            print(f"  ✗ {name}:")
            for d in diffs:
                print(f"      {d}")
        else:
            print(f"    {name}: matches")

    rp, gp = ref["perf"], gen["perf"]
    if rp or gp:
        print("\n-- Perf block --")
        if rp and not gp:
            print(f"  ✗ hand-made has perf block (top {_fmt(rp.get('top'))}, "
                  f"bottom {_fmt(rp.get('bottom'))}, date {_fmt(rp.get('date'))}); generated has none")
        elif rp and gp:
            for k in ("top", "bottom", "date"):
                mark = "  " if str(rp.get(k)) == str(gp.get(k)) else "✗ "
                print(f"  {mark}{k:<8} {_fmt(rp.get(k)):<20} {_fmt(gp.get(k)):<20}")

    if ref["casing"] or gen["casing"]:
        print("\n-- Casing table --")
        print(f"  hand-made rows: {len(ref['casing'])}, generated rows: {len(gen['casing'])}")
        if ref["casing"] and not gen["casing"]:
            for r in ref["casing"]:
                print(f"    hand-made only: {[ _fmt(v) for v in r if v is not None ]}")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    which = set(a.lower() for a in sys.argv[1:]) or {"south_moon", "reay"}

    if "south_moon" in which:
        print("Generating South Moon via WCR-PDF pipeline (Docling ~60s)...")
        path = generate_south_moon()
        diff(SOUTH_MOON_REF, path, "South Moon 5-31-32-C4-3H (4301353996) — PDF pipeline")

    if "reay" in which:
        print("\nGenerating Reay via From-Database pipeline...")
        path = generate_reay()
        diff(REAY_REF, path, "Reay 16-29-30-B4-2H (4301354722) — legacy DB pipeline")


if __name__ == "__main__":
    main()
