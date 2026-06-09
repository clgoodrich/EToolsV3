"""Batch compare: for every test APD that has BOTH an input PDF and an
original (hand-made) Casing Review Excel, run the real generate pipeline and
diff the generated workbook against the original.

Mirrors the Casing Review tab's generate() path (APD parse -> DB survey ->
promote/process -> clearance -> traversal + dx offsets -> svc.generate),
then recalculates both workbooks with LibreOffice (openpyxl can't evaluate
formulas) and compares the section sheets' meaningful cells.

Run: PYTHONPATH=. .venv/Scripts/python.exe scripts/compare_apd_batch.py
Writes a report to output/apd_compare_report.txt
"""
from __future__ import annotations

import re
import glob
import os
import subprocess
import traceback
from collections import defaultdict
from pathlib import Path

import openpyxl

from etools.core.pdf.apd_parser import parse_apd_pdf
from etools.core.casing_review.promote import (
    normalize_survey_dataframe,
    well_header_from_apd,
)
from etools.core.casing_review.sections import (
    apd_summary_footages,
    build_section_traversal,
    dx_survey_path_offsets,
    survey_kop_footages,
)
from etools.models.survey import SurveyFrame
from etools.repositories import SurveyRepository
from etools.services import CasingReviewService
from etools.services.survey_service import SurveyService
from etools.services.clearance_service import ClearanceService

SOFFICE = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")
SECTION_SHEETS = ["SHL Section"] + [f"BHL Section {i}" for i in range(1, 8)]
REPORT = Path("output/apd_compare_report.txt")
RECALC_DIR = Path("output/_cmp_recalc")

# Cells we compare on each section sheet. The reference-point block
# (Surface/KOP/Landing/TD) + the visible bearing grid (print area A22:M40).
REF_ROWS = {7: "Surface", 8: "KOP", 9: "Landing", 10: "TD"}
REF_COLS = {4: "MD", 5: "N-S", 6: "NSdir", 7: "E-W", 8: "EWdir", 9: "FNL/FSL", 11: "FEL/FWL"}


def api10(s: str) -> str | None:
    m = re.search(r"(\d{10,14})", s)
    return m.group(1)[:10] if m else None


def find_pairs() -> dict:
    pdfs = defaultdict(list)
    for p in glob.glob("**/application_*.pdf", recursive=True):
        a = api10(os.path.basename(p))
        if a:
            pdfs[a].append(p)
    xls = defaultdict(list)
    for p in glob.glob("**/Casing Review_*.xlsx", recursive=True):
        if p.startswith("output" + os.sep) or "_recalc" in p:
            continue
        a = api10(os.path.basename(p))
        if a:
            xls[a].append(p)

    def pick_pdf(paths):
        # Prefer a non-"Check" file, shallowest path.
        paths = sorted(paths, key=lambda p: ("check" in p.lower(), p.count(os.sep), len(p)))
        return paths[0]

    def pick_xls(paths):
        # Prefer the canonical tests/APD/ root file; avoid Error/ copies.
        def score(p):
            low = p.lower()
            return (
                "error" in low,
                "apd" + os.sep + "apd" in low,  # the weird nested dup
                low.count(os.sep),
                len(p),
            )
        return sorted(paths, key=score)[0]

    out = {}
    for a in sorted(set(pdfs) & set(xls)):
        out[a] = (pick_pdf(pdfs[a]), pick_xls(xls[a]))
    return out


def recalc(path: Path, tag: str) -> Path | None:
    """Recalc ``path`` into a per-side subdir.

    CRITICAL: the generated file and the original share the SAME filename
    (Casing Review_<api>_<name>.xlsx), so they MUST recalc into separate
    output dirs — otherwise LibreOffice (which names output by input stem)
    overwrites one with the other and we'd compare a file against itself.
    """
    if not SOFFICE.exists():
        return None
    outdir = RECALC_DIR / tag
    outdir.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.run(
            [str(SOFFICE), "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(outdir), str(path)],
            check=True, capture_output=True, timeout=180,
        )
    except Exception:
        return None
    cand = outdir / (path.stem + ".xlsx")
    return cand if cand.exists() else None


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _eq(a, b, *, tol=0.5) -> bool:
    if a is None and b is None:
        return True
    na, nb = _num(a), _num(b)
    if na is not None and nb is not None:
        return abs(na - nb) <= tol
    return str(a).strip().upper() == str(b).strip().upper()


def compare_sheet(gen_ws, orig_ws):
    """Return (ref_diffs, grid_match, grid_total, grid_diffs[:8])."""
    ref_diffs = []
    for r, rname in REF_ROWS.items():
        if r == 8:
            continue  # KOP intentionally differs (originals leave it default)
        for c, cname in REF_COLS.items():
            gv = gen_ws.cell(r, c).value
            ov = orig_ws.cell(r, c).value
            if not _eq(gv, ov):
                ref_diffs.append(f"{rname}.{cname}: gen={gv!r} orig={ov!r}")
    # Bearing grid print area A22:M40.
    total = match = 0
    gdiffs = []
    for row in range(22, 41):
        for col in range(1, 14):
            gv = gen_ws.cell(row, col).value
            ov = orig_ws.cell(row, col).value
            if gv in (None, "") and ov in (None, ""):
                continue
            total += 1
            if _eq(gv, ov, tol=0.5):
                match += 1
            elif len(gdiffs) < 8:
                coord = openpyxl.utils.get_column_letter(col) + str(row)
                gdiffs.append(f"{coord}: gen={gv!r} orig={ov!r}")
    return ref_diffs, match, total, gdiffs


def run_one(api, pdf, xls, log):
    log(f"\n{'='*70}\n{api}\n  PDF: {pdf}\n  ORIG: {xls}")
    apd = parse_apd_pdf(Path(pdf))
    log(f"  well={apd.well_name!r} locations={len(apd.locations)}")

    repo = SurveyRepository()
    results = repo.get_points_by_api_lateral(api[:10], "0000")
    chosen = next((c for c in ("AsDrilled", "Planned") if c in results and not results[c].empty), None)
    survey_df = results[chosen] if chosen else None
    log(f"  survey={chosen or 'NONE (synthetic fallback)'}")

    section_locations = dx = None
    if survey_df is not None:
        header = well_header_from_apd(apd)
        citing = header.citing_type or "Planned"
        surveys = {citing: normalize_survey_dataframe(survey_df)}
        processed = SurveyService().process([header], surveys)
        sr = processed[citing]
        ps = sr.frames[SurveyFrame.TRUE]
        clearance = ClearanceService().calculate(ps, kop_md=sr.kop.md, landing_md=sr.landing_md)
        crossings = build_section_traversal(apd.locations, clearance.points)
        section_locations = [c.to_location_row() for c in crossings] or None
        # Prefer the APD's stated kickoff MD over the survey-detected KOP.
        kop_md = apd.kop_md_ft if apd.kop_md_ft is not None else sr.kop.md
        dx = dx_survey_path_offsets(clearance.points, kop_md=kop_md, landing_md=sr.landing_md)
        log(f"  traversal={[c.conc for c in crossings]}")

    footages = list(apd_summary_footages(apd.locations) or [None, None, None])
    # Computed-KOP contingency: no "Location At Kickoff Point" in the APD ->
    # take the K.O. Point footages from the survey at the back-projected KOP.
    if footages and footages[0] is None and section_locations is not None:
        footages[0] = survey_kop_footages(clearance.points, sr.kop.md)
    result = CasingReviewService().generate(
        apd_data=apd, survey=survey_df, frac_gradient_override_psi_per_ft=1.0,
        section_locations=section_locations, dx_survey_locations=dx,
        dx_survey_footages=footages if any(footages) else None,
    )
    gen_path = result.output_path

    gen_rc = recalc(gen_path, "gen")
    orig_rc = recalc(Path(xls), "orig")
    if gen_rc is None or orig_rc is None:
        log("  !! recalc failed; cannot compare evaluated values")
        return
    gwb = openpyxl.load_workbook(gen_rc, data_only=True)
    owb = openpyxl.load_workbook(orig_rc, data_only=True)

    # FINAL (Total Depth) footage delta on the SHL summary — the value the
    # user cares about. Row 10, I = N/S footage, K = E/W footage.
    gshl = gwb["SHL Section"] if "SHL Section" in gwb.sheetnames else None
    oshl = owb["SHL Section"] if "SHL Section" in owb.sheetnames else None
    final = "n/a"
    if gshl is not None and oshl is not None:
        def _d(c):
            a, b = _num(gshl.cell(10, c).value), _num(oshl.cell(10, c).value)
            return abs(a - b) if (a is not None and b is not None) else None
        di, dk = _d(9), _d(11)
        parts = []
        parts.append(f"N/S d={di:.1f}ft" if di is not None else "N/S gen=" + repr(gshl.cell(10, 9).value))
        parts.append(f"E/W d={dk:.1f}ft" if dk is not None else "E/W gen=" + repr(gshl.cell(10, 11).value))
        final = " | ".join(parts)
    log(f"  >>> FINAL (TD) footage: {final}")

    tot_match = tot_cells = 0
    for sh in SECTION_SHEETS:
        if sh not in gwb.sheetnames or sh not in owb.sheetnames:
            continue
        ref_diffs, gmatch, gtot, gdiffs = compare_sheet(gwb[sh], owb[sh])
        tot_match += gmatch
        tot_cells += gtot
        pct = (100 * gmatch / gtot) if gtot else 100
        # Only print sheets that have something in the original (skip blanks).
        if gtot == 0 and not ref_diffs:
            continue
        log(f"  -- {sh}: grid {gmatch}/{gtot} ({pct:.0f}%) | ref-diffs={len(ref_diffs)}")
        for d in ref_diffs:
            if not d.startswith("KOP"):  # KOP intentionally differs (orig=default)
                log(f"        REF {d}")
        for d in gdiffs:
            log(f"        GRID {d}")
    overall = (100 * tot_match / tot_cells) if tot_cells else 0
    log(f"  >>> OVERALL grid match: {tot_match}/{tot_cells} ({overall:.0f}%)")
    return overall


def main():
    REPORT.parent.mkdir(exist_ok=True)
    lines = []

    def log(s):
        try:
            print(s)
        except UnicodeEncodeError:
            print(s.encode("ascii", "replace").decode("ascii"))
        lines.append(s)

    pairs = find_pairs()
    log(f"Matched APDs (PDF + original Excel): {len(pairs)}")
    summary = []
    for api, (pdf, xls) in pairs.items():
        try:
            overall = run_one(api, pdf, xls, log)
            summary.append((api, overall))
        except Exception as exc:
            log(f"  !! FAILED: {exc}")
            log("  " + traceback.format_exc().replace("\n", "\n  "))
            summary.append((api, None))
    log(f"\n{'='*70}\nSUMMARY")
    for api, ov in summary:
        s = f"{ov:.0f}%" if isinstance(ov, (int, float)) else "ERROR"
        log(f"  {api}: {s}")
    REPORT.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nReport written to {REPORT}")


if __name__ == "__main__":
    main()
