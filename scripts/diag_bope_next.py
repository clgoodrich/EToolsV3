"""Diagnostic: generate one well, recalc, dump TVDs / Next rows / BOPE sheet."""
from __future__ import annotations
import sys
from pathlib import Path
import openpyxl

from scripts.compare_apd_batch import find_pairs, recalc
from etools.core.pdf.apd_parser import parse_apd_pdf
from etools.core.casing_review.promote import normalize_survey_dataframe, well_header_from_apd
from etools.core.casing_review.sections import (
    apd_summary_footages, build_section_traversal, dx_survey_path_offsets, survey_kop_footages,
)
from etools.models.survey import SurveyFrame
from etools.repositories import SurveyRepository
from etools.services import CasingReviewService
from etools.services.survey_service import SurveyService
from etools.services.clearance_service import ClearanceService


def gen(api, pdf):
    apd = parse_apd_pdf(Path(pdf))
    repo = SurveyRepository()
    results = repo.get_points_by_api_lateral(api[:10], "0000")
    chosen = next((c for c in ("AsDrilled", "Planned") if c in results and not results[c].empty), None)
    survey_df = results[chosen] if chosen else None
    section_locations = dx = None
    clearance = sr = None
    if survey_df is not None:
        header = well_header_from_apd(apd)
        citing = header.citing_type or "Planned"
        surveys = {citing: normalize_survey_dataframe(survey_df)}
        processed = SurveyService().process([header], surveys)
        sr = processed[citing]
        ps = sr.frames[SurveyFrame.TRUE]
        clearance = ClearanceService().calculate(ps, kop_md=sr.kop.md, landing_md=sr.landing_md)
        crossings = build_section_traversal(apd.locations, clearance.points)
        section_locations = [c.to_location_row() for c in crossings] or None
        kop_md = apd.kop_md_ft if apd.kop_md_ft is not None else sr.kop.md
        dx = dx_survey_path_offsets(clearance.points, kop_md=kop_md, landing_md=sr.landing_md)
    footages = list(apd_summary_footages(apd.locations) or [None, None, None])
    if footages and footages[0] is None and section_locations is not None:
        footages[0] = survey_kop_footages(clearance.points, sr.kop.md)
    result = CasingReviewService().generate(
        apd_data=apd, survey=survey_df, frac_gradient_override_psi_per_ft=1.0,
        section_locations=section_locations, dx_survey_locations=dx,
        dx_survey_footages=footages if any(footages) else None,
    )
    return result.output_path, apd


def main():
    pairs = find_pairs()
    api = sys.argv[1] if len(sys.argv) > 1 else next(iter(pairs))
    pdf = pairs[api][0]
    print(f"WELL {api}\n PDF {pdf}")
    out, apd = gen(api, pdf)
    print(f" GEN {out}")
    print(f" apd.kop_md_ft={apd.kop_md_ft} locations={[loc.name for loc in apd.locations]}")
    rc = recalc(Path(out), "diag")
    if rc is None:
        print(" recalc FAILED")
        return
    wb = openpyxl.load_workbook(rc, data_only=True)
    cr = wb["Casing Review"]
    print("\n== Casing Review per-string TVD (B19/B34/B49/B64) & MD set depth ==")
    for s, (tvd_r, sd_r) in enumerate([(19,12),(34,27),(49,42),(64,57)], 1):
        print(f"  String {s}: TVD(B{tvd_r})={cr.cell(tvd_r,2).value!r}  SetDepthMD(D{sd_r})={cr.cell(sd_r,4).value!r}"
              f"  MW(B{tvd_r-1})={cr.cell(tvd_r-1,2).value!r}")
    print("\n== Next rows (24/39/54/69): B=NextSetDepth D=NextMW F=NextBHP I=FracInit L=MaxAnticShoe ==")
    for r in (24, 39, 54, 69):
        print(f"  row {r}: B={cr.cell(r,2).value!r} D={cr.cell(r,4).value!r} F={cr.cell(r,6).value!r} "
              f"I={cr.cell(r,9).value!r} L={cr.cell(r,12).value!r}")
    if "BOPE" in wb.sheetnames:
        b = wb["BOPE"]
        print("\n== BOPE sheet key cells ==")
        for coord in ["C5","D5","E5","C6","D6","E6","C7","D7","E7","C8","D8","E8",
                      "C9","D9","E9","C10","D10","E10","C11","F11",
                      "C14","C16","C17","C19","C20","C21","D16","D17","D19",
                      "C24","C34","C44"]:
            print(f"  {coord}={b[coord].value!r}")
    # also dump DxSurvey TVD table extent post-gen (formulas vs values)
    dgen = openpyxl.load_workbook(rc, data_only=True)["DxSurvey"]
    print("\n== DxSurvey survey table H (TVD) rows 16-30 ==")
    for r in range(16, 31):
        print(f"  r{r}: MD(B)={dgen.cell(r,2).value!r} TVD(H)={dgen.cell(r,8).value!r}")


if __name__ == "__main__":
    main()
