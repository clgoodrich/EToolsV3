"""End-to-end verify: APD -> promote -> generate, then check the four
reference points (Surface / KOP / Landing / TD) populate on the section
sheets for a problem well. Mirrors the Casing Review tab's generate() path.

Run: PYTHONPATH=. .venv/Scripts/python.exe scripts/verify_four_points.py 4304756010
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

from etools.core.pdf.apd_parser import parse_apd_pdf
from etools.core.casing_review.promote import (
    normalize_survey_dataframe,
    well_header_from_apd,
)
from etools.core.casing_review.sections import (
    build_section_traversal,
    dx_survey_path_offsets,
)
from etools.models.survey import SurveyFrame
from etools.repositories import SurveyRepository
from etools.services import CasingReviewService
from etools.services.survey_service import SurveyService
from etools.services.clearance_service import ClearanceService


def main(api: str) -> None:
    pdf = Path(f"application_{api}.pdf")
    print(f"== parsing {pdf} ==")
    apd = parse_apd_pdf(pdf)
    print(f"   well={apd.well_name!r} api={apd.api!r} locations={len(apd.locations)}")

    # Survey from DB (Planned), like _try_db_survey.
    repo = SurveyRepository()
    results = repo.get_points_by_api_lateral(api[:10], "0000")
    chosen = next((c for c in ("AsDrilled", "Planned") if c in results and not results[c].empty), None)
    if chosen is None:
        print("   !! no DB survey")
        return
    survey_df = results[chosen]
    print(f"   survey={chosen} stations={len(survey_df)}")

    # Promote: header + normalized survey.
    header = well_header_from_apd(apd)
    citing = header.citing_type or "Planned"
    surveys = {citing: normalize_survey_dataframe(survey_df)}

    # post_load: process + clearance.
    processed = SurveyService().process([header], surveys)
    sr = processed[citing]
    print(f"   KOP md={sr.kop.md} landing_md={sr.landing_md}")
    clr = ClearanceService()
    ps = sr.frames[SurveyFrame.TRUE]
    clearance = clr.calculate(ps, kop_md=sr.kop.md, landing_md=sr.landing_md)
    points = clearance.points

    # Build the same inputs generate() builds.
    crossings = build_section_traversal(apd.locations, points)
    section_locations = [c.to_location_row() for c in crossings] or None
    print(f"   traversal: {[c.conc for c in crossings]}")
    dx = dx_survey_path_offsets(points, kop_md=sr.kop.md, landing_md=sr.landing_md)
    print(f"   dx_survey rows (KOP/Landing/TD): {dx}")

    # Generate.
    result = CasingReviewService().generate(
        apd_data=apd,
        survey=survey_df,
        frac_gradient_override_psi_per_ft=1.0,
        section_locations=section_locations,
        dx_survey_locations=dx,
    )
    out = result.output_path
    print(f"== generated {out} ==")

    recalc = _libre_recalc(out)
    if recalc is None:
        print("!! LibreOffice recalc unavailable; skipping evaluated check")
        return
    wbv = openpyxl.load_workbook(recalc, data_only=True)
    names = ["SHL Section"] + [f"BHL Section {i}" for i in range(1, 8)]
    for sh in names:
        ws = wbv[sh]
        errs = [
            c.coordinate
            for row in ws.iter_rows(min_row=1, max_row=70, min_col=1, max_col=13)
            for c in row
            if isinstance(c.value, str) and c.value.startswith("#")
        ]
        print(f"-- {sh:16s} visible(A-M) errors={len(errs)} {errs[:5]}")
    shl = wbv["SHL Section"]
    print("-- SHL Section four points (evaluated) --")
    for r, name in ((7, "Surface"), (8, "KOP"), (9, "Landing"), (10, "TD")):
        print(
            f"   {name:8s} MD={shl.cell(r,4).value!r} "
            f"NS={shl.cell(r,5).value!r}{shl.cell(r,6).value or ''} "
            f"EW={shl.cell(r,7).value!r}{shl.cell(r,8).value or ''}"
        )


def _libre_recalc(path: Path):
    import subprocess
    soffice = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")
    if not soffice.exists():
        return None
    outdir = path.parent / "_recalc"
    outdir.mkdir(exist_ok=True)
    subprocess.run(
        [str(soffice), "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(path)],
        check=True, capture_output=True, timeout=120,
    )
    return outdir / path.name


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "4304756010")
