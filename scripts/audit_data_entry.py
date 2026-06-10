"""Data-entry completeness audit.

For each well that has a hand-made reference Casing Review workbook, generate
ours and diff every *human-entered* (literal, non-formula) value on the main
data-entry surfaces — Casing Review inputs (cols A-K + knobs) and the BOPE
sheet. Reports cells the original has filled that we leave BLANK (missing
automation) or fill DIFFERENTLY (wrong automation). Labels match in both and
drop out automatically.

Inputs are values (not formulas), so no LibreOffice recalc is needed.

Run: PYTHONPATH=. .venv/Scripts/python.exe scripts/audit_data_entry.py [N]
"""
from __future__ import annotations

import sys
from collections import defaultdict

import openpyxl

from scripts.compare_apd_batch import find_pairs
from scripts.diag_bope_next import gen


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _same(a, b) -> bool:
    if a is None and b is None:
        return True
    na, nb = _num(a), _num(b)
    if na is not None and nb is not None:
        return abs(na - nb) <= 0.6
    return str(a).strip().upper() == str(b).strip().upper()


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


# (sheet, max_row, max_col) — the data-entry surfaces.
SURFACES = [("Casing Review", 69, 11), ("BOPE", 51, 13)]


def audit_one(api, pdf):
    out, _apd = gen(api, pdf)
    genwb = openpyxl.load_workbook(out, data_only=False)
    # original sits next to its API in find_pairs
    orig_path = pairs[api][1]
    owb = openpyxl.load_workbook(orig_path, data_only=False)
    missing, wrong = [], []
    for sheet, mr, mc in SURFACES:
        if sheet not in owb.sheetnames or sheet not in genwb.sheetnames:
            continue
        ows, gws = owb[sheet], genwb[sheet]
        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                ov = ows.cell(r, c).value
                if ov is None or _is_formula(ov):
                    continue  # only literals the human typed
                gv = gws.cell(r, c).value
                if _same(ov, gv):
                    continue
                coord = f"{sheet}!{openpyxl.utils.get_column_letter(c)}{r}"
                if gv is None or (isinstance(gv, str) and gv.strip() == ""):
                    missing.append((coord, ov))
                elif not _is_formula(gv):
                    wrong.append((coord, gv, ov))
    return missing, wrong


def main():
    global pairs
    pairs = find_pairs()
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 4
    # representative sample: a 4301353701, two Butcher Butte, a Ute Shavanaugh
    picks = ["4301353701", "4301353749", "4301353774", "4304756010"][:n]
    agg_missing = defaultdict(int)
    agg_wrong = defaultdict(int)
    for api in picks:
        if api not in pairs:
            continue
        miss, wrong = audit_one(api, pairs[api][0])
        print(f"\n{'='*64}\n{api}  missing={len(miss)} wrong={len(wrong)}")
        for coord, ov in miss[:40]:
            print(f"   MISSING {coord:22} orig={ov!r}")
            agg_missing[coord.split('!')[0] + '!' + ''.join(ch for ch in coord.split('!')[1] if ch.isalpha())] += 1
        for coord, gv, ov in wrong[:40]:
            print(f"   WRONG   {coord:22} ours={gv!r}  orig={ov!r}")
            agg_wrong[coord.split('!')[0] + '!' + ''.join(ch for ch in coord.split('!')[1] if ch.isalpha())] += 1
    print(f"\n{'='*64}\nAGGREGATE missing-by-column:")
    for k, v in sorted(agg_missing.items(), key=lambda kv: -kv[1]):
        print(f"   {k:20} {v}")
    print("AGGREGATE wrong-by-column:")
    for k, v in sorted(agg_wrong.items(), key=lambda kv: -kv[1]):
        print(f"   {k:20} {v}")


if __name__ == "__main__":
    main()
