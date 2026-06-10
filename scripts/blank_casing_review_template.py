"""Clear the operator-input cells in the Casing Review template.

The template ships as a copy of a real Newfield workbook so it carries
the original well's values in every STRING block. Run this once to wipe
those cells so the generator starts from a clean slate. Formulas, the
Casing Strengths lookup, defined names, and reference sheets are left
intact.

Usage:
    .venv/Scripts/python scripts/blank_casing_review_template.py
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
TEMPLATE = REPO / "etools" / "core" / "casing_review" / "templates" / "casing_review_template.xlsx"

INPUT_CELLS_HEADER = ("B4", "B5", "B6", "B9")
# Per-STRING-block input cells, relative to block-top row (10 / 25 / 40 / 55).
# Data row (top+2): B, C, D, E, F, G, H, I, J, K, L, M, N, O
# Engineering knobs: B(top+7..top+13)
PER_BLOCK_DATA_COLS = ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O")
PER_BLOCK_KNOB_ROWS = (7, 8, 10, 11, 12, 13)  # B17/B18/B20/B21/B22/B23 for STRING 1


def main() -> None:
    if not TEMPLATE.exists():
        print(f"Template not found: {TEMPLATE}", file=sys.stderr)
        sys.exit(1)

    # Back the original up before we touch it.
    backup = TEMPLATE.with_suffix(".xlsx.bak")
    if not backup.exists():
        shutil.copyfile(TEMPLATE, backup)
        print(f"Backed up original -> {backup.name}")

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["Casing Review"]

    cleared = 0
    for addr in INPUT_CELLS_HEADER:
        if ws[addr].value is not None and not _is_formula(ws[addr].value):
            ws[addr] = None
            cleared += 1

    for top in (10, 25, 40, 55):
        data_row = top + 2
        for col in PER_BLOCK_DATA_COLS:
            addr = f"{col}{data_row}"
            if ws[addr].value is not None and not _is_formula(ws[addr].value):
                ws[addr] = None
                cleared += 1
        for off in PER_BLOCK_KNOB_ROWS:
            addr = f"B{top + off}"
            if ws[addr].value is not None and not _is_formula(ws[addr].value):
                ws[addr] = None
                cleared += 1

    wb.save(TEMPLATE)
    print(f"Cleared {cleared} cells in {TEMPLATE.name}")


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


if __name__ == "__main__":
    main()
