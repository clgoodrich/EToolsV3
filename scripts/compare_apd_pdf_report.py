"""Generate a visual PDF comparing eTools-generated Casing Reviews against the
hand-made legacy Excel originals — for a couple of representative wells plus
the obviously aberrant ones.

Reuses the batch harness (scripts/compare_apd_batch.py) to run the real
generate pipeline and recalc both workbooks with LibreOffice, then renders a
side-by-side reference-point comparison table per well into a single PDF.

Run: PYTHONPATH=. .venv/Scripts/python.exe scripts/compare_apd_pdf_report.py
Writes: output/casing_review_comparison.pdf
"""
from __future__ import annotations

from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl

from scripts.compare_apd_batch import find_pairs, run_one

# A couple of clean/representative wells + the known-aberrant ones.
GOOD = ["4301353749", "4301353764"]          # <=5 ft both axes — clean match
ABERRANT = ["4301353784", "4301353786", "4304756010"]  # Error-folder outliers
SELECTED = GOOD + ABERRANT

OUT_PDF = Path("output/casing_review_comparison.pdf")

REF_ROWS = [(7, "Surface"), (8, "K.O. Point"), (9, "Prod. Interval"), (10, "Total Depth")]
# col -> header
REF_COLS = [
    (4, "MD"),
    (5, "N-S off"),
    (6, "dir"),
    (7, "E-W off"),
    (8, "dir"),
    (9, "FNL/FSL"),
    (10, "code"),
    (11, "FEL/FWL"),
    (12, "code"),
]


def _fmt(v):
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        return f"{f:,.1f}" if f % 1 else f"{int(f):,}"
    except (TypeError, ValueError):
        return str(v)


def _cells(ws):
    """Return rows of [label, *col-values] for the reference block."""
    rows = []
    for r, label in REF_ROWS:
        vals = [label] + [_fmt(ws.cell(r, c).value) for c, _ in REF_COLS]
        rows.append(vals)
    return rows


def _diff_mask(gen_rows, orig_rows, tol=5.0):
    """Bool per cell: True if gen vs orig differ meaningfully (numeric tol=5ft)."""
    mask = []
    for gr, orr in zip(gen_rows, orig_rows):
        mrow = [False]  # label col never flagged
        for gv, ov in zip(gr[1:], orr[1:]):
            try:
                differ = abs(float(gv.replace(",", "")) - float(ov.replace(",", ""))) > tol
            except (ValueError, AttributeError):
                differ = str(gv).strip().upper() != str(ov).strip().upper()
            mrow.append(differ)
        mask.append(mrow)
    return mask


def _table(ax, title, rows, mask, color):
    ax.set_title(title, fontsize=9, fontweight="bold", loc="left", color=color)
    ax.axis("off")
    headers = ["Point"] + [h for _, h in REF_COLS]
    tbl = ax.table(cellText=rows, colLabels=headers, loc="center", cellLoc="center")
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(6.5)
    tbl.scale(1, 1.4)
    ncols = len(headers)
    for (r, c), cell in tbl.get_cells().items() if hasattr(tbl, "get_cells") else tbl._cells.items():
        cell.set_edgecolor("#cccccc")
        if r == 0:
            cell.set_facecolor("#222222")
            cell.set_text_props(color="white", fontweight="bold")
        elif r - 1 < len(mask) and c < ncols and mask[r - 1][c]:
            cell.set_facecolor("#ffd6d6")  # differing cell


def build_page(pdf, api, kind, gen_ws, orig_ws):
    gen_rows = _cells(gen_ws)
    orig_rows = _cells(orig_ws)
    mask = _diff_mask(gen_rows, orig_rows)

    fig = plt.figure(figsize=(11, 8.5))
    fig.suptitle(
        f"Casing Review reference points — API {api}  ({kind})",
        fontsize=12, fontweight="bold",
    )
    ax_g = fig.add_axes([0.04, 0.55, 0.92, 0.32])
    ax_o = fig.add_axes([0.04, 0.14, 0.92, 0.32])
    _table(ax_g, "eTools GENERATED", gen_rows, mask, "#1a6e1a")
    _table(ax_o, "LEGACY ORIGINAL (hand-made)", orig_rows, mask, "#9a2b2b")

    # Footnote: final-footage delta on the Total Depth row.
    def _d(c):
        try:
            return abs(float(gen_ws.cell(10, c).value) - float(orig_ws.cell(10, c).value))
        except (TypeError, ValueError):
            return None
    di, dk = _d(9), _d(11)
    note = (
        "Pink cells differ by >5 ft (or text mismatch).  "
        f"Total-Depth footage delta:  N/S = {di:.1f} ft" if di is not None else "N/S n/a"
    )
    note += (f"  |  E/W = {dk:.1f} ft" if dk is not None else "  |  E/W n/a")
    fig.text(0.04, 0.06, note, fontsize=8, color="#444444")
    pdf.savefig(fig)
    plt.close(fig)


def main():
    OUT_PDF.parent.mkdir(exist_ok=True)
    pairs = find_pairs()

    def log(s):
        print(s)

    with PdfPages(OUT_PDF) as pdf:
        # Cover page.
        fig = plt.figure(figsize=(11, 8.5))
        fig.text(0.5, 0.7, "Casing Review — Generated vs. Legacy", ha="center", fontsize=20, fontweight="bold")
        fig.text(0.5, 0.62, "eTools output compared against hand-made Excel originals", ha="center", fontsize=12, color="#555555")
        body = (
            "Representative wells (clean match, <=5 ft both axes):\n"
            "    4301353749, 4301353764\n\n"
            "Aberrant wells (originals live in tests/APD/Error/ and disagree\n"
            "with their own APD permits; eTools matches the permit):\n"
            "    4301353784  (E/W ~66 ft),  4301353786  (E/W ~33 ft),\n"
            "    4304756010  (cross-township excursion, E/W ~170 ft)\n\n"
            "Each page shows the four reference points (Surface / KOP / Landing /\n"
            "Total Depth) from the eTools-generated workbook above and the legacy\n"
            "original below. Pink = cells differing by >5 ft.\n\n"
            "NOTE on K.O. Point: when the APD prints a kickoff ('KOP: <md>' MD'\n"
            "+ a 'Location At Kickoff Point' row), eTools now reads it straight\n"
            "from the permit. On the 4304756xxx wells the legacy originals are\n"
            "WRONG here (they copied the surface footages with MD ~1,100 and zero\n"
            "offsets) -- so pink on the KOP row means eTools matches the document\n"
            "and the hand-made original does not."
        )
        fig.text(0.12, 0.45, body, ha="left", va="top", fontsize=10, family="monospace")
        pdf.savefig(fig)
        plt.close(fig)

        for api in SELECTED:
            if api not in pairs:
                log(f"  !! {api} not in matched pairs; skipping")
                continue
            pdf_path, xls_path = pairs[api]
            kind = "clean match" if api in GOOD else "ABERRANT (Error-folder original)"
            log(f"\n### {api} ({kind})")
            # run_one regenerates + recalcs into output/_cmp_recalc/{gen,orig}.
            try:
                run_one(api, pdf_path, xls_path, log)
            except Exception as exc:
                log(f"  !! pipeline failed: {exc}")
                continue
            # Generated stem differs from the original's, so find the gen-dir
            # xlsx for this run by matching the API instead.
            gen_dir = Path("output/_cmp_recalc/gen")
            gen_files = sorted(gen_dir.glob(f"*{api}*.xlsx"))
            orig_rc = Path("output/_cmp_recalc/orig") / (Path(xls_path).stem + ".xlsx")
            if not gen_files or not orig_rc.exists():
                log(f"  !! recalc outputs missing for {api}; skipping page")
                continue
            gwb = openpyxl.load_workbook(gen_files[-1], data_only=True)
            owb = openpyxl.load_workbook(orig_rc, data_only=True)
            gws = gwb["SHL Section"] if "SHL Section" in gwb.sheetnames else None
            ows = owb["SHL Section"] if "SHL Section" in owb.sheetnames else None
            if gws is None or ows is None:
                log(f"  !! SHL Section missing for {api}; skipping page")
                continue
            build_page(pdf, api, kind, gws, ows)

    print(f"\nPDF written to {OUT_PDF.resolve()}")


if __name__ == "__main__":
    main()
