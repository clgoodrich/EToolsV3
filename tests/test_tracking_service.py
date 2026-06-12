"""TrackingWCR.xlsx updater — create, append, and update-in-place by API."""
from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook

from etools.services.tracking_service import update_tracking_workbook


def test_create_and_update_row(tmp_path) -> None:
    path = tmp_path / "TrackingWCR.xlsx"
    update_tracking_workbook(
        path=path,
        api="4301354722",
        well_name="Reay 16-29-30-B4-2H",
        operator="Javelin",
        sundry_no=131236,
        date_filed=datetime(2026, 6, 1),
        returns=1,
        comp_sum=True,
        logs_included=True,
        edits=["utms", "footages"],
    )
    ws = load_workbook(path).active
    assert ws.cell(1, 5).value == "API"
    assert ws.cell(2, 5).value == 4301354722
    assert ws.cell(2, 4).value == 131236
    assert ws.cell(2, 10).value == "y"   # comp sum
    assert ws.cell(2, 11).value == "n"   # drilling sum
    assert ws.cell(2, 16).value == "y"   # edited (edits present)
    assert ws.cell(2, 17).value == "utms/footages"
    assert isinstance(ws.cell(2, 1).value, int)  # days average

    # Second well appends; same API updates in place.
    update_tracking_workbook(path=path, api="4301353996", well_name="South Moon",
                             operator="Javelin", sundry_no=None, date_filed=None)
    update_tracking_workbook(path=path, api="4301354722", well_name="Reay",
                             operator="Javelin", sundry_no=99, date_filed=None, returns=2)
    ws = load_workbook(path).active
    assert ws.max_row == 3
    assert ws.cell(2, 4).value == 99 and ws.cell(2, 3).value == 2
    assert ws.cell(3, 5).value == 4301353996
    assert ws.cell(3, 2).value is None  # no filed date
