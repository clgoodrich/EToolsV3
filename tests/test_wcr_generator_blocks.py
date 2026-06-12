"""WCR Excel optional blocks — perf header (E1:G2) + casing table (row 16+)."""
from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from etools.core.wcr import generate_wcr_excel
from etools.models import WCRLocationRow, WCRWellInfo
from etools.services.wcr_service import _parse_label, _perf_summary


def _info() -> WCRWellInfo:
    return WCRWellInfo(api_well_no="43013547220000", well_name="Test Well")


def _rows() -> list[WCRLocationRow]:
    return [WCRLocationRow(name="SHL", measured_depth=0, tvd=0, easting=555200, northing=4458447)]


def _casing_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Feature": "Hole", "Top_MD": 0, "Bottom_MD": 128, "Diameter": 245},
            {
                "Feature": "Surface Casing", "Top_MD": 0, "Bottom_MD": 3282,
                "Diameter": 9.62, "Weight": 40, "Grade": "API Grade J-55 Casing",
                "ConnectionType": "BTC", "CementTop": 0, "CementBottom": 3282,
                "CementType": "Class G Cement", "Sacks": 345, "Yield": 1.43,
                "CementWeight": 14.5,
            },
        ]
    )


def test_perf_and_casing_blocks(tmp_path) -> None:
    out = tmp_path / "wcr.xlsx"
    generate_wcr_excel(
        info=_info(),
        location_rows=_rows(),
        output_path=out,
        perf_top_md=10283,
        perf_bottom_md=20704,
        perf_date="10/27/2024",
        casing=_casing_df(),
    )
    ws = load_workbook(out)["Sheet"]
    assert [ws.cell(1, c).value for c in (5, 6, 7)] == ["Perf Top", "Perf Bottom", "Perf Date"]
    assert [ws.cell(2, c).value for c in (5, 6, 7)] == [10283, 20704, "10/27/2024"]
    assert ws.cell(16, 1).value == "Feature"
    assert ws.cell(16, 13).value == "Cement Weight"
    assert [ws.cell(17, c).value for c in (1, 2, 3, 4)] == ["Hole", 0, 128, 245]
    assert ws.cell(18, 6).value == "API Grade J-55 Casing"
    assert ws.cell(18, 12).value == 1.43


def test_blocks_omitted_when_no_data(tmp_path) -> None:
    out = tmp_path / "wcr.xlsx"
    generate_wcr_excel(info=_info(), location_rows=_rows(), output_path=out)
    ws = load_workbook(out)["Sheet"]
    assert ws.cell(1, 5).value is None
    assert ws.cell(16, 1).value is None


def test_parse_label() -> None:
    assert _parse_label("14 2S 5W U") == {
        "section": "14", "township": "2", "township_dir": "S",
        "range": "5", "range_dir": "W", "baseline": "U",
    }
    assert _parse_label(None) == {}
    assert _parse_label("garbage") == {}


def test_perf_summary() -> None:
    perfs = pd.DataFrame(
        {
            "Top_MD": [10283.0, 12000.0],
            "Bottom_MD": [11000.0, 20704.0],
            "PerfDate": ["2024-10-01", "2024-10-27"],
        }
    )
    top, bottom, date = _perf_summary(perfs)
    assert top == 10283 and bottom == 20704
    assert date == "10/27/2024"
    assert _perf_summary(pd.DataFrame()) == (None, None, None)
