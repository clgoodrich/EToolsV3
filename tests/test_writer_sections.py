"""Writer regression: the Casing Review section sheets are formula-driven.

The template is designed so that ONLY the ``SHL Section`` sheet takes real
PLSS inputs; every BHL Section sheet derives its inputs by reference to the
SHL sheet and auto-detects its own crossed section from the survey path
stored in ``DxSurvey`` rows 8-10. An earlier writer hardcoded each BHL
sheet's N7/I7/K7/L38 — which broke the template's adjacency arithmetic
(``"" + number`` → ``#VALUE!`` cascading into every bearing cell) and left
the BHL bearing grids blank.

These tests lock the corrected contract:
    * SHL sheet gets real PLSS inputs.
    * BHL sheets keep their native template formulas (NOT overwritten).
    * The DxSurvey path-offset rows that drive section detection are filled.

Uses ``plat_repo=None`` so the tests are hermetic (no DB).
"""

from __future__ import annotations

import openpyxl
import pandas as pd

from etools.core.casing_review.domain import CasingDesign
from etools.core.casing_review.sections import (
    build_section_traversal,
    dx_survey_path_offsets,
)
from etools.core.casing_review.writer import write_casing_review
from etools.models import APDLocationRow


def _loc(name: str, sec: str, **foot) -> APDLocationRow:
    return APDLocationRow(
        name=name, section=sec, township="3", township_dir="S",
        range="2", range_dir="W", meridian="U", **foot,
    )


def _loc31e(name: str, sec: str, **foot) -> APDLocationRow:
    """A location in T3S R1E U (the Ute Shavanaugh township)."""
    return APDLocationRow(
        name=name, section=sec, township="3", township_dir="S",
        range="1", range_dir="E", meridian="U", **foot,
    )


def _design() -> CasingDesign:
    return CasingDesign(
        company="ACME", well_name="TEST 16-23", api="4301399999",
        frac_gradient_psi_per_ft=1.0, strings=[],
    )


def test_only_shl_sheet_gets_real_inputs(tmp_path) -> None:
    """Each crossed section gets its full PLSS written to a sequential
    sheet so the bearing-grid DGET resolves that exact section."""
    locs = [
        _loc("Location at Surface", "23", fnl=660, fel=1980),
        _loc("Top of Uppermost Producing Zone", "26", fnl=300, fel=900),
        _loc("At Total Depth", "35", fsl=250, fel=800),
    ]
    pts = pd.DataFrame(
        [
            {"Conc": "2303S02WU", "FNL": 660, "FSL": None, "FEL": 1980, "FWL": None},
            {"Conc": "2403S02WU", "FNL": 1320, "FSL": None, "FEL": 50, "FWL": None},
            {"Conc": "2603S02WU", "FNL": 300, "FSL": None, "FEL": 900, "FWL": None},
            {"Conc": "3503S02WU", "FNL": None, "FSL": 250, "FEL": 800, "FWL": None},
        ]
    )
    section_locations = [c.to_location_row() for c in build_section_traversal(locs, pts)]

    out = tmp_path / "cr.xlsx"
    write_casing_review(
        _design(), out, section_locations=section_locations, plat_repo=None
    )
    wb = openpyxl.load_workbook(out)

    # SHL sheet: real inputs; P7/R7 as INT codes (SHL criteria read ints).
    shl = wb["SHL Section"]
    assert shl["N7"].value == 23
    assert shl["O7"].value == 3
    assert shl["P7"].value == 2      # S → int code on the SHL sheet
    assert shl["R7"].value == 2      # W → int code

    # BHL sheets: full PLSS written. N7 = real section, L38 = real section,
    # P7/R7 as the "S"/"W" STRINGS (BHL criteria read =IF($P$7="S",…)).
    # Traversal is 23 (SHL), 24, 26, 35.
    expected = {"BHL Section 1": 24, "BHL Section 2": 26, "BHL Section 3": 35}
    for sheet, sec in expected.items():
        ws = wb[sheet]
        assert ws["N7"].value == sec, f"{sheet} N7={ws['N7'].value!r}"
        assert ws["L38"].value == sec, f"{sheet} L38={ws['L38'].value!r}"
        assert ws["O7"].value == 3
        assert ws["P7"].value == "S"
        assert ws["R7"].value == "W"

    # Well/API header stamped on every section sheet.
    for sheet in ("SHL Section", "BHL Section 1", "BHL Section 2", "BHL Section 3"):
        assert wb[sheet]["C2"].value == "TEST 16-23"
        assert wb[sheet]["C3"].value == "4301399999"


def test_eight_section_sheets_always_present_and_visible(tmp_path) -> None:
    """SHL + BHL 1-7 are all provisioned and left visible, regardless of how
    many the wellbore actually uses (none are hidden)."""
    locs = [
        _loc("Location at Surface", "23", fnl=660, fel=1980),
        _loc("At Total Depth", "24", fsl=250, fel=800),
    ]
    pts = pd.DataFrame(
        [
            {"Conc": "2303S02WU", "FNL": 660, "FSL": None, "FEL": 1980, "FWL": None},
            {"Conc": "2403S02WU", "FNL": 1320, "FSL": None, "FEL": 50, "FWL": None},
        ]
    )
    section_locations = [c.to_location_row() for c in build_section_traversal(locs, pts)]
    assert len(section_locations) == 2  # SHL(23) + BHL(24)

    out = tmp_path / "cr.xlsx"
    write_casing_review(
        _design(), out, section_locations=section_locations, plat_repo=None
    )
    wb = openpyxl.load_workbook(out)
    expected_sheets = ["SHL Section"] + [f"BHL Section {i}" for i in range(1, 8)]
    for name in expected_sheets:
        assert name in wb.sheetnames, f"missing {name}"
        assert wb[name].sheet_state == "visible", f"{name} should be visible"
    # The crossed sections are written; unused sheets stay in template state.
    assert wb["BHL Section 1"]["L38"].value == 24


def test_unused_section_sheets_are_blanked_but_visible(tmp_path) -> None:
    """Sheets the wellbore never reaches stay present and visible, but their
    bearing-grid formulas are cleared so they don't show a #VALUE! grid."""
    locs = [
        _loc("Location at Surface", "23", fnl=660, fel=1980),
        _loc("At Total Depth", "24", fsl=250, fel=800),
    ]
    pts = pd.DataFrame(
        [
            {"Conc": "2303S02WU", "FNL": 660, "FSL": None, "FEL": 1980, "FWL": None},
            {"Conc": "2403S02WU", "FNL": 1320, "FSL": None, "FEL": 50, "FWL": None},
        ]
    )
    section_locations = [c.to_location_row() for c in build_section_traversal(locs, pts)]
    assert len(section_locations) == 2  # SHL(23) + BHL 1(24); BHL 2-7 unused

    out = tmp_path / "cr.xlsx"
    write_casing_review(
        _design(), out, section_locations=section_locations, plat_repo=None
    )
    wb = openpyxl.load_workbook(out)

    # Used sheets keep their bearing-grid formulas.
    used = wb["BHL Section 1"]
    assert any(
        isinstance(c.value, str) and c.value.startswith("=")
        for row in used.iter_rows(min_row=15, max_row=63, max_col=13)
        for c in row
    )

    # Unused sheets: visible, header intact, but no grid formulas remain.
    for name in ("BHL Section 2", "BHL Section 7"):
        ws = wb[name]
        assert ws.sheet_state == "visible", f"{name} should stay visible"
        assert ws["C2"].value == "TEST 16-23"  # header preserved
        leftover = [
            c.coordinate
            for row in ws.iter_rows(min_row=15, max_row=63, max_col=13)
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ]
        assert not leftover, f"{name} still has grid formulas: {leftover[:5]}"


def test_cross_township_disables_continuation(tmp_path) -> None:
    """When the wellbore threads a section in a different township than the
    surface, the cross-section continuation is peeled off the BHL sheets so
    each section shows its own bearings (no #VALUE! from the broken walk)."""
    # SHL in T3S R1E sec 5; detour into T2S R1E sec 32; back into sec 8.
    locs = [_loc31e("Location at Surface", "5"), _loc31e("At Total Depth", "8")]
    pts = pd.DataFrame(
        [
            {"Conc": "0503S01EU", "measured_depth": 0.0, "FNL": 100, "FEL": 100},
            {"Conc": "3202S01EU", "measured_depth": 6000.0, "FNL": 50, "FEL": 50},
            {"Conc": "0803S01EU", "measured_depth": 14000.0, "FSL": 250, "FEL": 800},
        ]
    )
    section_locations = [c.to_location_row() for c in build_section_traversal(locs, pts)]
    assert [loc.section for loc in section_locations] == ["5", "32", "8"]

    out = tmp_path / "cr.xlsx"
    write_casing_review(
        _design(), out, section_locations=section_locations, plat_repo=None
    )
    wb = openpyxl.load_workbook(out)
    # Continuation (an IF referencing 'SHL Section'!$BE$) must be gone from
    # the BHL sheets — peeled down to each section's own value.
    for sheet in ("BHL Section 1", "BHL Section 2"):
        for coord in ("E17", "K17", "B30", "W38"):
            v = wb[sheet][coord].value
            assert not (isinstance(v, str) and "'SHL Section'!$BE$" in v), (
                f"{sheet}!{coord} still has continuation: {v!r}"
            )


def test_dx_survey_offsets_are_written(tmp_path) -> None:
    """The KOP/Prod/TD path offsets land in DxSurvey C8:E10 — these drive
    the BHL sheets' section detection."""
    locs = [_loc("Location at Surface", "23", fnl=660, fel=1980)]
    pts = pd.DataFrame(
        [
            {"Conc": "2303S02WU", "measured_depth": 0.0, "n_offset": 0.0, "e_offset": 0.0},
            {"Conc": "2303S02WU", "measured_depth": 5000.0, "n_offset": -200.0, "e_offset": 900.0},
            {"Conc": "2403S02WU", "measured_depth": 9000.0, "n_offset": -800.0, "e_offset": 1500.0},
            {"Conc": "2403S02WU", "measured_depth": 14000.0, "n_offset": -5000.0, "e_offset": 1700.0},
        ]
    )
    section_locations = [c.to_location_row() for c in build_section_traversal(locs, pts)]
    dx = dx_survey_path_offsets(pts, kop_md=5000.0, landing_md=9000.0)

    out = tmp_path / "cr.xlsx"
    write_casing_review(
        _design(), out, section_locations=section_locations,
        dx_survey_locations=dx, plat_repo=None,
    )
    wb = openpyxl.load_workbook(out)
    dxs = wb["DxSurvey"]
    # Row 8 = KOP (md 5000), row 9 = Prod (md 9000), row 10 = TD (md 14000).
    assert dxs["C8"].value == 5000.0
    assert dxs["D8"].value == -200.0
    assert dxs["E8"].value == 900.0
    assert dxs["C9"].value == 9000.0
    assert dxs["C10"].value == 14000.0
    assert dxs["D10"].value == -5000.0


def test_dx_survey_path_offsets_picks_nearest_md() -> None:
    """dx_survey_path_offsets returns (md, n, e) at the station nearest each
    reference MD, defaulting TD to the deepest station."""
    pts = pd.DataFrame(
        [
            {"measured_depth": 0.0, "n_offset": 0.0, "e_offset": 0.0},
            {"measured_depth": 100.0, "n_offset": -1.0, "e_offset": 5.0},
            {"measured_depth": 200.0, "n_offset": -9.0, "e_offset": 9.0},
        ]
    )
    rows = dx_survey_path_offsets(pts, kop_md=90.0, landing_md=None)
    assert rows[0] == (100.0, -1.0, 5.0)     # nearest to 90 is the 100 station
    assert rows[1] is None                   # no landing MD
    assert rows[2] == (200.0, -9.0, 9.0)     # TD defaults to deepest

    # Missing offset columns → empty (writer falls back to template default).
    assert dx_survey_path_offsets(pd.DataFrame({"measured_depth": [1, 2]})) == []
    assert dx_survey_path_offsets(None) == []
