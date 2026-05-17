"""
Well Completion Report (WCR) generator for EToolsV3.

This module generates WCR Excel reports from survey and clearance data.
"""

from typing import Optional
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font

from data.models.survey import ProcessedSurvey
from data.models.clearance import ClearanceResults, SectionFootage
from data.models.well import WellInfo
from config.logging_config import get_logger
from config.settings import settings
from utils.errors import ExcelGenerationError

logger = get_logger(__name__)


class WCRGenerator:
    """Generates Well Completion Report Excel files."""

    def __init__(self, template_path: Optional[Path] = None):
        """
        Initialize WCR generator.

        Args:
            template_path: Path to WCR Excel template. If None, looks in project root.
        """
        self.logger = logger

        if template_path is None:
            template_path = settings.paths.root_dir / 'WCR_Empty.xlsm'

        self.template_path = template_path

        if not self.template_path.exists():
            self.logger.warning(f"WCR template not found at {template_path}")

    def generate_wcr(
        self,
        well_info: WellInfo,
        survey: ProcessedSurvey,
        clearances: ClearanceResults,
        footage: SectionFootage,
        output_path: Optional[Path] = None
    ) -> Path:
        """
        Generate WCR Excel file.

        Args:
            well_info: Well information.
            survey: Processed survey data.
            clearances: Clearance results.
            footage: Section footage calculations.
            output_path: Output file path. If None, generates default name.

        Returns:
            Path to generated WCR file.

        Raises:
            ExcelGenerationError: If generation fails.
        """
        try:
            self.logger.info(
                f"Generating WCR for API={well_info.api_number}, "
                f"Lateral={survey.header.lateral_name}"
            )

            # Generate output filename if not provided
            if output_path is None:
                filename = f"{survey.header.lateral_name}_{well_info.api_number}_WCR.xlsx"
                output_path = settings.paths.output_dir / filename

            # Load template or create new workbook
            if self.template_path.exists():
                wb = load_workbook(self.template_path, keep_vba=True)
            else:
                wb = openpyxl.Workbook()

            # Get or create sheets
            if 'WCR' in wb.sheetnames:
                ws_main = wb['WCR']
            else:
                ws_main = wb.active
                ws_main.title = 'WCR'

            # Fill in well information
            self._fill_well_info(ws_main, well_info, survey)

            # Fill in survey data
            self._fill_survey_data(ws_main, survey)

            # Fill in clearance data
            self._fill_clearance_data(ws_main, clearances)

            # Fill in footage data
            self._fill_footage_data(ws_main, footage)

            # Save workbook
            wb.save(output_path)

            self.logger.info(f"WCR saved to {output_path}")

            return output_path

        except Exception as e:
            self.logger.error(f"WCR generation failed: {e}", exc_info=True)
            raise ExcelGenerationError(
                f"Failed to generate WCR: {str(e)}",
                details={'api': well_info.api_number}
            )

    def _fill_well_info(
        self,
        worksheet,
        well_info: WellInfo,
        survey: ProcessedSurvey
    ):
        """
        Fill well information section.

        Args:
            worksheet: Excel worksheet.
            well_info: Well information.
            survey: Survey data.
        """
        # Common cell mappings (adjust based on actual template)
        cell_mappings = {
            'B2': well_info.api_number,
            'B3': well_info.well_name or '',
            'B4': well_info.operator_name or survey.header.operator_name or '',
            'B5': survey.header.lateral_name,
            'B6': datetime.now().strftime('%Y-%m-%d')
        }

        for cell, value in cell_mappings.items():
            try:
                worksheet[cell] = value
            except Exception as e:
                self.logger.warning(f"Failed to set cell {cell}: {e}")

    def _fill_survey_data(self, worksheet, survey: ProcessedSurvey):
        """
        Fill survey data section.

        Args:
            worksheet: Excel worksheet.
            survey: Processed survey.
        """
        # Starting row for survey data (adjust based on template)
        start_row = 10

        survey_data = survey.data[[
            'measured_depth', 'inclination', 'azimuth',
            'tvd', 'northing', 'easting'
        ]].copy()

        # Write headers
        headers = ['MD', 'INC', 'AZ', 'TVD', 'Northing', 'Easting']
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        # Write data
        for row_idx, (_, row) in enumerate(survey_data.iterrows(), start=start_row + 1):
            for col_idx, col in enumerate(survey_data.columns, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=row[col])

    def _fill_clearance_data(self, worksheet, clearances: ClearanceResults):
        """
        Fill clearance data section.

        Args:
            worksheet: Excel worksheet.
            clearances: Clearance results.
        """
        # This would fill clearance data into appropriate cells
        # Implementation depends on template structure
        pass

    def _fill_footage_data(self, worksheet, footage: SectionFootage):
        """
        Fill section footage data.

        Args:
            worksheet: Excel worksheet.
            footage: Section footage calculations.
        """
        # Starting row for footage data
        start_row = 100

        # Write headers
        headers = ['Section', 'Total Footage', 'Entry MD', 'Exit MD']
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        # Write data
        for row_idx, section_footage in enumerate(footage.section_footages, start=start_row + 1):
            worksheet.cell(row=row_idx, column=1, value=section_footage.section_label)
            worksheet.cell(row=row_idx, column=2, value=section_footage.total_footage)
            worksheet.cell(row=row_idx, column=3, value=section_footage.measured_depth_in)
            worksheet.cell(row=row_idx, column=4, value=section_footage.measured_depth_out)

    def create_summary_dataframe(
        self,
        survey: ProcessedSurvey,
        clearances: ClearanceResults,
        footage: SectionFootage
    ) -> pd.DataFrame:
        """
        Create summary DataFrame for WCR.

        Args:
            survey: Processed survey.
            clearances: Clearance results.
            footage: Section footage.

        Returns:
            Summary DataFrame.
        """
        summary_data = {
            'API Number': survey.header.api_number,
            'Lateral': survey.header.lateral_name,
            'Total MD': survey.total_md,
            'Max TVD': survey.max_tvd,
            'Max Inclination': survey.max_inclination,
            'KOP MD': survey.kop_md,
            'Sections Penetrated': clearances.num_sections,
            'Min FNL': clearances.min_fnl,
            'Min FSL': clearances.min_fsl,
            'Min FEL': clearances.min_fel,
            'Min FWL': clearances.min_fwl,
            'Total Drilled': footage.total_drilled
        }

        return pd.DataFrame([summary_data])
