"""WCR stress test: edge cases + generate Excel + compare against existing."""
import sys, io, time, pathlib, traceback
sys.path.insert(0, str(pathlib.Path(__file__).parent / "app"))

import pandas as pd
import numpy as np
import openpyxl

out = io.open("wcr_stress_report.txt", "w", encoding="utf-8")

def log(msg=""):
    print(msg)
    out.write(msg + "\n")
    out.flush()

WCR_DIR = pathlib.Path(r"C:\ETools_Portable\WCR")
CG_DIR = WCR_DIR / "CG Files"
OUTPUT_DIR = pathlib.Path(r"C:\ETools_Portable\app\output")

# ─── Imports ────────────────────────────────────────────────────────────────
from etools.core.pdf.wcr_parser import parse_wcr_pdf
from etools.services.wcr_pdf_service import WCRPdfService

log("=" * 70)
log("WCR STRESS TEST REPORT")
log("=" * 70)
log()

# ─── SECTION 1: Edge case parsing ───────────────────────────────────────────
log("## 1. WCR Parse Edge Cases")
log()

edge_cases = [
    ("Non-existent file", r"C:\ETools_Portable\WCR\does_not_exist.pdf", {"mode": "rules", "skip_docling": True}),
    ("Empty file", r"C:\ETools_Portable\STRESS_TEST_LOG.md", {"mode": "rules", "skip_docling": True}),  # text not pdf
    ("APD file as WCR", r"C:\ETools_Portable\APD\application_43013537270000.pdf", {"mode": "rules", "skip_docling": True}),
    ("Template PDF", r"C:\ETools_Portable\WCR\Template_pdf_test.pdf", {"mode": "rules", "skip_docling": True}),
    ("Complete PDF", r"C:\ETools_Portable\WCR\complete.pdf", {"mode": "rules", "skip_docling": True}),
    ("Complete (1) PDF", r"C:\ETools_Portable\WCR\complete (1).pdf", {"mode": "rules", "skip_docling": True}),
    ("complete (2) PDF", r"C:\ETools_Portable\WCR\complete (2).pdf", {"mode": "rules", "skip_docling": True}),
    ("max_pages=1", r"C:\ETools_Portable\WCR\WCR 43013539950000.pdf", {"mode": "rules", "skip_docling": True, "max_pages": 1}),
    ("max_pages=5", r"C:\ETools_Portable\WCR\WCR 43013539950000.pdf", {"mode": "rules", "skip_docling": True, "max_pages": 5}),
    ("mode=llm (no ollama)", r"C:\ETools_Portable\WCR\WCR 43013539950000.pdf", {"mode": "llm", "skip_docling": True}),
]

for name, path, kwargs in edge_cases:
    try:
        t0 = time.time()
        data = parse_wcr_pdf(path, **kwargs)
        elapsed = time.time() - t0
        log(f"  PASS [{name}]: api={data.api!r} well={data.well_name!r} casing={len(data.casing)} "
            f"formations={len(data.formations)} stages={len(data.perf_stages)} "
            f"positions={len(data.positions)} warnings={data.warnings} ({elapsed:.2f}s)")
    except Exception as e:
        elapsed = time.time() - t0
        log(f"  RAISE [{name}]: {type(e).__name__}: {e} ({elapsed:.2f}s)")

log()

# ─── SECTION 2: Parse all root WCR PDFs ─────────────────────────────────────
log("## 2. Parse All Root WCR PDFs")
log()

wcr_pdfs = list(WCR_DIR.glob("*.pdf"))
for pdf in sorted(wcr_pdfs):
    try:
        t0 = time.time()
        data = parse_wcr_pdf(str(pdf), mode="rules", skip_docling=True)
        elapsed = time.time() - t0
        # Check for data quality issues
        issues = []
        if not data.api:
            issues.append("NO API")
        if not data.well_name:
            issues.append("NO WELL NAME")
        if not data.total_md_ft:
            issues.append("NO TOTAL MD")
        if not data.positions:
            issues.append("NO POSITIONS")
        if data.perf_stages:
            sizes = [s.size_in for s in data.perf_stages if s.size_in]
            max_size = max(sizes) if sizes else None
            if max_size and max_size > 1.5:
                issues.append(f"SUSPICIOUS PERF SIZE: {max_size}")
            counts = [s.num_perfs for s in data.perf_stages if s.num_perfs]
            max_perfs = max(counts) if counts else None
            if max_perfs and max_perfs > 1000:
                issues.append(f"SUSPICIOUS PERF COUNT: {max_perfs}")
        log(f"  {pdf.name}: api={data.api}, well={data.well_name!r}, "
            f"md={data.total_md_ft}, casing={len(data.casing)}, "
            f"fmns={len(data.formations)}, stages={len(data.perf_stages)}, "
            f"{'ISSUES: ' + str(issues) if issues else 'OK'} ({elapsed:.2f}s)")
    except Exception as e:
        log(f"  ERROR {pdf.name}: {type(e).__name__}: {e}")
        log(traceback.format_exc())

log()

# ─── SECTION 3: WCR Excel generation + comparison ────────────────────────────
log("## 3. WCR Excel Generate + Compare vs Existing")
log()

# Matched pairs: (wcr_pdf, existing_excel)
# WCR requires a survey; use a minimal vertical well survey so the service can run.
# (Empty survey raises ValueError: Cannot process an empty survey.)
# Synthetic vertical survey: straight down to ~22,000 ft then horizontal lateral.
def make_vertical_survey(td_ft: float) -> pd.DataFrame:
    """Build a simple survey that goes vertical then kicks to ~90 deg for a lateral."""
    rows = [
        {"MeasuredDepth": 0.0,          "Inclination": 0.0,  "Azimuth": 0.0},
        {"MeasuredDepth": td_ft * 0.5,  "Inclination": 5.0,  "Azimuth": 0.0},
        {"MeasuredDepth": td_ft * 0.7,  "Inclination": 45.0, "Azimuth": 0.0},
        {"MeasuredDepth": td_ft * 0.8,  "Inclination": 85.0, "Azimuth": 0.0},
        {"MeasuredDepth": td_ft,        "Inclination": 90.0, "Azimuth": 0.0},
    ]
    return pd.DataFrame(rows)

EMPTY_SURVEY = pd.DataFrame(columns=["MeasuredDepth", "Inclination", "Azimuth"])

PAIRS = [
    (
        WCR_DIR / "WCR 43013539960000.pdf",
        CG_DIR / "South_Moon_5-31-32-C4-3H_4301353996_WCR.xlsx",
    ),
    (
        WCR_DIR / "WCR 43013540190000.pdf",
        CG_DIR / "Federal_3-18D-65_4301354019_WCR.xlsx",
    ),
    (
        WCR_DIR / "WCR 43013540300000.pdf",
        CG_DIR / "Federal_15-7D-65_4301354030_WCR.xlsx",
    ),
]


def colnum_to_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def load_sheet_values(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    data = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                data[(cell.row, cell.column)] = cell.value
    return data


def compare_sheets(wb_gen, wb_exist, sheet_name, max_diffs=30):
    gen_data = load_sheet_values(wb_gen, sheet_name)
    exist_data = load_sheet_values(wb_exist, sheet_name)
    diffs = []
    all_keys = set(gen_data.keys()) | set(exist_data.keys())
    for (r, c) in sorted(all_keys):
        gval = gen_data.get((r, c))
        eval_ = exist_data.get((r, c))
        if gval == eval_:
            continue
        try:
            gf = float(gval) if gval is not None else None
            ef = float(eval_) if eval_ is not None else None
            if gf is not None and ef is not None:
                if abs(gf - ef) < 1e-6:
                    continue
                pct = abs(gf - ef) / max(abs(ef), 1e-9) * 100
                diffs.append({"row": r, "col": c, "generated": gval, "existing": eval_,
                              "type": "numeric", "pct_diff": pct})
                continue
        except (TypeError, ValueError):
            pass
        gs = str(gval).strip() if gval is not None else ""
        es = str(eval_).strip() if eval_ is not None else ""
        if gs != es:
            diffs.append({"row": r, "col": c, "generated": gval, "existing": eval_,
                          "type": "string", "pct_diff": None})
    return diffs


def generate_wcr_excel(wcr_pdf_path):
    """Generate WCR Excel using a synthetic survey. Pre-parse with skip_docling."""
    svc = WCRPdfService()
    wcr_data = parse_wcr_pdf(str(wcr_pdf_path), mode="rules", skip_docling=True)
    td = wcr_data.total_md_ft or 20000.0
    survey = make_vertical_survey(td)
    result = svc.generate(
        wcr_data=wcr_data,
        surveys=survey,
    )
    return pathlib.Path(result.output_path)


svc = WCRPdfService()
for wcr_path, exist_path in PAIRS:
    log("=" * 70)
    log(f"WCR PDF: {wcr_path.name}")
    log(f"Existing: {exist_path.name}")
    log("=" * 70)

    if not exist_path.exists():
        log("  ERROR: existing Excel not found — skipping")
        continue

    try:
        log("  Generating WCR Excel (no survey)...")
        t0 = time.time()
        gen_path = generate_wcr_excel(wcr_path)
        elapsed = time.time() - t0
        log(f"  Generated: {gen_path} ({elapsed:.1f}s)")
    except Exception as e:
        log(f"  ERROR generating: {type(e).__name__}: {e}")
        log(traceback.format_exc())
        log()
        continue

    # Compare
    log("  Loading workbooks for comparison...")
    wb_gen = openpyxl.load_workbook(str(gen_path), data_only=True)
    wb_exist = openpyxl.load_workbook(str(exist_path), data_only=True)

    gen_sheets = set(wb_gen.sheetnames)
    exist_sheets = set(wb_exist.sheetnames)
    only_in_gen = gen_sheets - exist_sheets
    only_in_exist = exist_sheets - gen_sheets
    common = gen_sheets & exist_sheets

    if only_in_gen:
        log(f"  Sheets only in GENERATED: {sorted(only_in_gen)}")
    if only_in_exist:
        log(f"  Sheets only in EXISTING: {sorted(only_in_exist)}")

    total_diffs = 0
    for sheet in sorted(common):
        diffs = compare_sheets(wb_gen, wb_exist, sheet)
        if not diffs:
            log(f"  [{sheet}] -- MATCH")
            continue
        log(f"\n  [{sheet}] -- {len(diffs)} differences")
        for d in diffs[:30]:
            cell = f"{colnum_to_letter(d['col'])}{d['row']}"
            if d["type"] == "numeric":
                log(f"    {cell}: GEN={d['generated']!r}  EXIST={d['existing']!r}  ({d['pct_diff']:.2f}%)")
            else:
                g = repr(str(d['generated'])[:60]) if d['generated'] is not None else 'None'
                e = repr(str(d['existing'])[:60]) if d['existing'] is not None else 'None'
                log(f"    {cell}: GEN={g}  EXIST={e}")
        if len(diffs) > 30:
            log(f"    ... ({len(diffs) - 30} more)")
        total_diffs += len(diffs)

    log(f"\n  TOTAL DIFFERENCES: {total_diffs}")
    log()

# ─── SECTION 4: WCR service edge cases ───────────────────────────────────────
log("## 4. WCR Service Edge Cases")
log()

svc = WCRPdfService()
# Pre-parse once to get TD for survey generation
_wcr_test_data = parse_wcr_pdf(str(WCR_DIR / "WCR 43013539950000.pdf"), mode="rules", skip_docling=True)
_td = _wcr_test_data.total_md_ft or 20000.0
GOOD_SURVEY = make_vertical_survey(_td)
TEST_WCR = str(WCR_DIR / "WCR 43013539950000.pdf")

edge_gen = [
    ("No args",              dict(surveys=EMPTY_SURVEY)),
    ("Empty survey DF",      dict(wcr_data=_wcr_test_data, surveys=EMPTY_SURVEY)),
    ("NaN surface lat",      dict(wcr_data=_wcr_test_data, surveys=GOOD_SURVEY,
                                  surface_lat=float('nan'), surface_lon=float('nan'))),
    ("None surface lat",     dict(wcr_data=_wcr_test_data, surveys=GOOD_SURVEY,
                                  surface_lat=None, surface_lon=None)),
    ("Extreme lat 91",       dict(wcr_data=_wcr_test_data, surveys=GOOD_SURVEY,
                                  surface_lat=91.0, surface_lon=-109.0)),
    ("Survey with NaN",      dict(wcr_data=_wcr_test_data,
                                  surveys=pd.DataFrame({"MeasuredDepth": [0, float('nan'), 1000],
                                                        "Inclination":   [0, float('nan'), 5],
                                                        "Azimuth":        [0, float('nan'), 90]}))),
    ("Zero elevation",       dict(wcr_data=_wcr_test_data, surveys=GOOD_SURVEY,
                                  surface_elevation_ft=0.0)),
    ("Negative elevation",   dict(wcr_data=_wcr_test_data, surveys=GOOD_SURVEY,
                                  surface_elevation_ft=-500.0)),
    ("Good run (baseline)",  dict(wcr_data=_wcr_test_data, surveys=GOOD_SURVEY)),
]

for name, kwargs in edge_gen:
    try:
        t0 = time.time()
        result = svc.generate(**kwargs)
        elapsed = time.time() - t0
        log(f"  PASS [{name}]: output={result.output_path} ({elapsed:.1f}s)")
    except Exception as e:
        elapsed = time.time() - t0
        log(f"  RAISE [{name}]: {type(e).__name__}: {str(e)[:100]} ({elapsed:.2f}s)")

log()
log("WCR stress test complete.")
out.close()
print("Report written to wcr_stress_report.txt")
