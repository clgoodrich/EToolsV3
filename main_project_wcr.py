"""
Well Completion Report (WCR) Module

This module handles the generation and management of Well Completion Reports for oil and gas directional wells.
It interfaces with SQL databases to retrieve well information, processes survey data, and generates Excel reports
following regulatory requirements for well completion documentation.

The WCR (Well Completion Report) is a critical regulatory document that must be filed after completing
a well, containing details about the well's construction, survey data, casing information, and perforation zones.
"""

from typing import Dict, List, Tuple, Optional, Any, Union
import collections.abc
from datetime import date, datetime

# Handle deprecated collections aliases for compatibility
collections.Iterable = collections.abc.Iterable
collections.Mapping = collections.abc.Mapping
collections.MutableSet = collections.abc.MutableSet
collections.MutableMapping = collections.abc.MutableMapping

# Third-party imports
import numpy as np
import pandas as pd
import utm
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from PyQt5.QtGui import QPixmap, QStandardItem, QStandardItemModel
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidgetItem, QWidget, QVBoxLayout,
    QPushButton, QRadioButton, QButtonGroup, QMessageBox, QLabel,
    QScrollArea, QHeaderView, QAbstractItemView
)



class WCR_Main:
    """
    Main class for handling Well Completion Report (WCR) generation and processing.

    This class manages the entire WCR workflow including:
    - Retrieving well data from SQL databases
    - Processing directional survey information
    - Calculating UTM coordinates and footage from section lines
    - Generating regulatory-compliant Excel reports
    - Managing UI interactions for WCR data entry and validation

    The WCR is a mandatory regulatory filing that documents the as-drilled well path,
    casing design, and completion details for state regulatory compliance.
    """

    def __init__(
            self,
            df: Dict[str, pd.DataFrame],
            ui: Any,
            db: Any,
            loc_df: pd.DataFrame,
            spec_surveys: Dict[str, Any],
            north_ref: str
    ) -> None:
        """
        Initialize the WCR_Main class with well data and UI components.

        Args:
            df: Dictionary containing directional survey dataframes with 'drl' prefix
            ui: PyQt5 UI object containing all interface elements
            db: Database connection object for SQL queries
            loc_df: DataFrame containing well location data
            spec_surveys: Dictionary of special survey points (KOP, EOC, etc.)
            north_ref: North reference type ('T' for True North, 'G' for Grid North)

        The initialization sets up:
        - UI event connections for real-time data processing
        - Data models for table displays
        - Initial form states and values
        - Perforation interval data retrieval
        """
        # Initialize instance variables for data storage
        self.perf_date: Optional[str] = None
        self.perf_mods: Optional[List[float]] = None
        self.wcr_result_df: Optional[pd.DataFrame] = None

        # Filter dataframes to only include drilled ('drl') survey data
        filtered_dict = {k: v for k, v in df.items() if "drl" in k}
        self._sundries_df: Optional[pd.DataFrame] = None

        # Store initialization parameters
        self.df = filtered_dict
        self.spec_surveys = {k: v for k, v in spec_surveys.items() if "drl" in k}
        self.ui = ui
        self.db = db
        self.loc_df = loc_df
        self.north_ref = north_ref

        # Extract well identification information from UI
        self.api: str = self.ui.well_api_val.text()
        self.lateral: str = self.ui.lateral_name_line_edit.text()

        # Connect UI signals for real-time interpolation calculations
        # These connections enable automatic calculation of interpolated coordinates
        # when users input survey station data
        self.ui.prev_depth_line.textChanged.connect(self.interpolation_table_process)
        self.ui.prev_ns_line.textChanged.connect(self.interpolation_table_process)
        self.ui.prev_ew_line.textChanged.connect(self.interpolation_table_process)
        self.ui.cur_depth_line.textChanged.connect(self.interpolation_table_process)
        self.ui.next_depth_line.textChanged.connect(self.interpolation_table_process)
        self.ui.next_ns_line.textChanged.connect(self.interpolation_table_process)
        self.ui.next_ew_line.textChanged.connect(self.interpolation_table_process)

        # Initialize data models for table displays
        self.report_box_model = QStandardItemModel()
        self.ui.logs_report_box.setModel(self.report_box_model)
        self.ui.logs_report_box.setEditTriggers(QAbstractItemView.NoEditTriggers)

        # Connect button signals to their respective handler methods
        self.ui.find_submitted_logs_button.pressed.connect(self.create_well_logs_list)
        self.ui.find_submitted_sundries_button.pressed.connect(self.create_sundries_list)
        self.ui.sundries_combo_box.activated.connect(self.sundries_combo_box_process)
        self.ui.load_plat_button.pressed.connect(self.present_plat_data)
        self.ui.collect_survey_push.pressed.connect(self.process_survey_data)
        self.ui.update_personal_wcr.pressed.connect(self.update_personal_excel)
        self.ui.wcr_survey_northref_buttons.buttonClicked.connect(self.process_survey_data)
        self.ui.pushButton.pressed.connect(self.generate_excel_file)

        # Initialize coordinate display model
        self.coords_wcr_model = QStandardItemModel()
        self.ui.display_table_utm_locs.setModel(self.coords_wcr_model)
        self.ui.display_table_utm_locs.setEditTriggers(QAbstractItemView.NoEditTriggers)

        # Clear combo boxes and reset form state
        self.ui.survey_north_ref_combo.clear()
        self.ui.plat_north_ref_combo.clear()
        self.ui.sundries_combo_box.clear()
        self.ui.sundries_report_box.setText('')
        self.reset_check_boxes()

        # Retrieve and display perforation interval data
        perf_mods, self.perf_date = self.wcr_perf_mods()
        self.ui.perf_interval_top.setText(str(perf_mods[0]))
        self.ui.perf_interval_bottom.setText(str(perf_mods[1]))

        # Set north reference based on stored preference
        one_letter_id = self.north_ref[0].upper()
        if one_letter_id == 'T':
            self.ui.wcr_survey_northref_buttons.button(-2).setChecked(True)
        elif one_letter_id == 'G':
            self.ui.wcr_survey_northref_buttons.button(-3).setChecked(True)

    def reset_check_boxes(self) -> None:
        """
        Reset all UI checkboxes to unchecked state and clear text fields.

        This method ensures a clean state for the WCR form, preventing
        data carry-over between different well reports. It's called during
        initialization and when switching between wells.
        """
        # Reset all action taken checkboxes
        self.ui.action_taken_checkbox.setChecked(False)
        self.ui.comp_sum_checkbox.setChecked(False)
        self.ui.drill_sum_checkbox.setChecked(False)
        self.ui.cement_log_checkbox.setChecked(False)
        self.ui.logs_included_checkbox.setChecked(False)
        self.ui.asdrilled_excel_checkbox.setChecked(False)
        self.ui.action_taken__utms_checkbox.setChecked(False)
        self.ui.action_taken_footages_checkbox.setChecked(False)
        self.ui.action_taken_perfs_checkbox.setChecked(False)
        self.ui.action_taken_depths_checkbox.setChecked(False)
        self.ui.action_taken_other_lineedit.setText("")
        self.ui.action_taken_other_checkbox.setChecked(False)

    def savedPopup(self) -> None:
        """
        Display a simple confirmation dialog indicating data has been saved.

        This provides user feedback after successful Excel generation or
        database updates, ensuring users know their actions completed successfully.
        """
        choice = QMessageBox.warning(None, "Attention", "Data saved!", QMessageBox.Ok)

    def get_wcr_info(self) -> pd.DataFrame:
        """
        Retrieve WCR-specific well information from the database.

        Returns:
            DataFrame containing well header information including:
            - Operator name
            - Well name/number
            - API number
            - APD (Application for Permit to Drill) information
            - Construction details

        This query joins multiple tables to gather all required regulatory
        information for the WCR header section.
        """
        query = f"""select wi.*, dsh.OperatorName, dsh.WellNameNumber, dsh.APINumber, LateralName as ConstructKey
        from tblAPDWCRWellInfo wi
        join [dbo].tblAPD ta on wi.APDNo = ta.APDNo
        JOIN DirectionalSurveyHeader dsh ON LEFT(ta.API_WellNo, 10) = dsh.APINumber 
        where ta.API_WellNo = '{self.api}{self.lateral}' and dsh.LateralName = '{self.lateral}'"""

        wcr_info = self.db.query_to_dataframe(query)
        wcr_info = wcr_info.drop_duplicates(keep="first")
        return wcr_info

    def get_wcr_casing(self) -> pd.DataFrame:
        """
        Retrieve casing and cementing information for the well.

        Returns:
            DataFrame containing casing strings with:
            - Casing type (Surface, Intermediate, Production)
            - Depths (top and bottom)
            - Physical properties (diameter, weight, grade)
            - Cement details (top, bottom, type, volume)

        The data is sorted by casing type in drilling order to match
        regulatory reporting requirements.
        """
        query = f"""select Feature, [Top], Bottom, Diam, [Weight], Grade,
       [Connection Type], [Cement Top], [Cement Bottom], [Cement Type],
       Sacks, Yield, [Cement Weight]
        from well w
        inner join construct c on w.pkey=c.WellKey
        inner join vwDM_ConstructCasingCement vw on c.PKey = vw.PKey
        WHERE w.WellID = {self.api} and [Feature] is not Null"""

        # Define standard casing order for regulatory compliance
        custom_order = ['Hole', 'Surface Casing', 'Intermediate Casing',
                        'Production Casing', 'Production Casing 2', 'Tubing']

        wcr_casing = self.db.query_to_dataframe(query)

        # Filter to only include standard casing types
        wcr_casing = wcr_casing[wcr_casing['Feature'].isin(custom_order)]
        wcr_casing = wcr_casing.drop_duplicates(keep="first")

        # Sort by drilling sequence for proper report formatting
        wcr_casing['Feature'] = pd.Categorical(wcr_casing['Feature'],
                                               categories=custom_order,
                                               ordered=True)
        wcr_casing = wcr_casing.sort_values(['Feature', 'Bottom']).reset_index(drop=True)

        return wcr_casing

    def generate_excel_file(self) -> None:
        """
        Generate the complete WCR Excel file with all required data.

        This method orchestrates the entire Excel generation process:
        1. Retrieves well information and casing data
        2. Checks if survey data needs processing
        3. Calls the Excel creation method

        The generated Excel file follows state regulatory formatting requirements
        for well completion reports.
        """
        # Retrieve required data from database
        wcr_info = self.get_wcr_info()
        wcr_casing = self.get_wcr_casing()

        # Check if coordinate table has data
        model = self.ui.display_table_utm_locs.model()
        test = [[model.data(model.index(row, column)) for column in range(model.columnCount())]
                for row in range(model.rowCount())]

        # Process survey data if not already done
        if not test:
            self.process_survey_data()
            self.run_excel_process(wcr_info, wcr_casing)
        else:
            self.run_excel_process(wcr_info, wcr_casing)

    def run_excel_process(self, wcr_info: pd.DataFrame, wcr_casing: pd.DataFrame) -> None:
        """
        Create and format the WCR Excel file with all required sections.

        Args:
            wcr_info: DataFrame containing well header information
            wcr_casing: DataFrame containing casing and cement data

        The Excel file includes:
        - Well identification header
        - Survey coordinate data at key points
        - Casing and cementing details
        - Perforation intervals

        File naming follows the pattern: WellName_API_WCR.xlsx
        """
        # Initialize Excel formatting
        bold_font = Font(bold=True)
        wcr_info['ModifyDate'] = pd.to_datetime(wcr_info['ModifyDate'],
                                                format='%m/%d/%Y',
                                                errors='coerce')

        # Define header labels for well information section
        labels = ['WellName', 'API', 'Operator', 'ConstructKey' 'WellType',
                  'SpudDate', 'RotaryRigDate', 'TDReachedDate', 'CompletedOrAbandonedDate']

        # Define column headers for survey data
        survey_scale = ['measured_depth', 'tvd', 'easting', 'northing',
                        'FNL', 'FSL', 'FEL', 'FWL', 'Section', 'Township',
                        'Township_Direction', 'Range', 'Range_Direction', 'Baseline']

        # Define survey point labels
        survey_depths = ['SHL', 'Control_Point', 'Frac_Start', 'Frac_End', 'BHL']
        survey_casing = wcr_casing.columns

        # Extract survey data for Excel
        used_wcr = self.wcr_result_df[survey_scale]

        # Excel cell references
        label_cells = ['1', '2', '3', '4', '5', '6', '7', '8']
        dx_data_header_cells = 'ABCDEFGHIJKLMNOPQRSTUV'
        max_row = 0

        # Create new workbook
        wb = Workbook()
        sheet = wb.active

        # Set column widths for readability
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 15

        # Process date fields
        list_dates = ['SpudRigDate', 'RotaryRigDate', 'TDReachedDate', 'CompletedOrAbandonedDate']
        wcr_info[list_dates] = (
            wcr_info[list_dates]
            .apply(pd.to_datetime, errors='coerce')
            .fillna(pd.Timestamp('1900-01-01'))
        )

        # Extract spud date for special handling
        first_spud: pd.Timestamp = wcr_info['SpudRigDate'].iloc[0]
        spud_str = first_spud.strftime('%Y-%m-%d')

        # Convert dates to string format
        wcr_info[list_dates] = wcr_info[list_dates].apply(lambda s: s.dt.strftime('%Y-%m-%d'))
        print(wcr_info)
        # Prepare well information values
        wcr_info['CompletedOrAbandonedDate'] = pd.to_datetime(wcr_info['CompletedOrAbandonedDate'], errors='coerce')
        formatted_date = wcr_info['CompletedOrAbandonedDate'].dt.strftime('%Y-%m-%d').iloc[0]

        info_vals = [
            wcr_info['WellNameNumber'].iloc[0],
            wcr_info['APINumber'].iloc[0],
            wcr_info['OperatorName'].iloc[0],
            wcr_info['ConstructKey'].iloc[0],
            wcr_info['WellType'].iloc[0],
            wcr_info['SpudRigDate'].iloc[0],
            wcr_info['RotaryRigDate'].iloc[0],
            wcr_info['TDReachedDate'].iloc[0],
            wcr_info['CompletedOrAbandonedDate'].dt.strftime('%Y-%m-%d').iloc[0]
        ]

        # Write well information to Excel
        for i in range(len(labels)):
            sheet["A" + label_cells[i]] = labels[i]
            sheet["B" + label_cells[i]] = info_vals[i]

        # Write survey column headers
        for i in range(len(survey_scale)):
            sheet[dx_data_header_cells[i + 1] + "10"] = survey_scale[i]

        # Round footage values to whole numbers
        used_wcr[['FNL', 'FSL', 'FEL', 'FWL']] = used_wcr[['FNL', 'FSL', 'FEL', 'FWL']].round(0)
        used_wcr = used_wcr.drop_duplicates(keep='first')

        # Write survey data rows
        for idx, row in used_wcr.iterrows():
            sheet['A' + str(11 + idx)] = survey_depths[idx]
            max_row = 11 + idx
            for i in range(len(survey_scale)):
                sheet[dx_data_header_cells[i + 1] + str(max_row)] = row[i]

        # Write casing column headers
        for i in range(len(survey_casing)):
            sheet[dx_data_header_cells[i] + str(max_row + 2)] = survey_casing[i]

        # Write casing data
        for idx, row in wcr_casing.iterrows():
            for i in range(len(row)):
                sheet[dx_data_header_cells[i] + str(max_row + 3 + idx)] = row[i]

        # Add perforation data section
        sheet['E1'] = 'Perf Top'
        sheet['F1'] = 'Perf Bottom'
        sheet['G1'] = 'Perf Date'
        sheet['E1'].font = bold_font
        sheet['F1'].font = bold_font
        sheet['G1'].font = bold_font

        sheet['E2'] = self.perf_mods[0]
        sheet['F2'] = self.perf_mods[1]
        sheet['G2'] = self.perf_date

        # Apply bold formatting to headers
        for cell in sheet[10]:
            cell.font = bold_font
        for cell in sheet[max_row + 2]:
            cell.font = bold_font
        for cell in sheet['A']:
            cell.font = bold_font

        # Generate filename and save
        wb_name = "{}_{}_WCR.xlsx".format(wcr_info['WellNameNumber'].iloc[0],
                                          wcr_info['APINumber'].iloc[0])
        wb_name = wb_name.replace(" ", "_").replace("/", "_")
        wb.save(wb_name)
        self.savedPopup()

    def update_personal_excel(self) -> None:
        """
        Update personal tracking spreadsheet with WCR processing information.

        This method maintains a separate Excel file for tracking:
        - Processing dates and turnaround times
        - Document completeness checks
        - Actions taken during review
        - Return/revision counts

        The tracking helps monitor regulatory compliance and processing efficiency.
        """

        def get_all_edits() -> str:
            """
            Compile a summary string of all edits made during WCR processing.

            Returns:
                Slash-delimited string of edit types
            """
            utms_str = "utms" if self.ui.action_taken__utms_checkbox.isChecked() else ""
            footages_str = "footages" if self.ui.action_taken_footages_checkbox.isChecked() else ""
            perfs_str = "perfs" if self.ui.action_taken_perfs_checkbox.isChecked() else ""
            depths_str = "depths" if self.ui.action_taken_depths_checkbox.isChecked() else ""
            other_str = self.ui.action_taken_other_lineedit.text() if self.ui.action_taken_other_checkbox.isChecked() else ""
            return "/".join(filter(None, [utms_str, footages_str, perfs_str, depths_str, other_str]))

        # Open tracking spreadsheet
        file = r"TrackingWCR.xlsx"
        data = self.get_wcr_person_db_update()
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        # Find existing entry or create new row
        values = [cell.value for cell in sheet['E']]
        try:
            api_number = int(float(data['APINumber'].iloc[0]))
        except IndexError:
            api_number = int(float(self.api))

        if api_number in values:
            api_index = values.index(api_number) + 1
            max_row = api_index
        else:
            max_row = sheet.max_row + 1

        # Extract tracking data
        date_filed = data['SubmitDate'].iloc[0]
        counted_returns = self.ui.no_returns_box.text()
        ret_count = 0 if counted_returns == "" else counted_returns
        sundry_number = data['SundryNo'].iloc[0]
        well_name = data['WellNameNumber'].iloc[0]
        date_processed = date.today()
        now_object = datetime.now()
        current_time = now_object.time()

        date_as_datetime = datetime.combine(date_processed, current_time)
        days_average = int(abs(date_filed - date_as_datetime).days)
        company = data['OperatorName'].iloc[0]

        # Compile checkbox states
        action_taken = "y" if self.ui.action_taken_checkbox.isChecked() else "n"
        comp_sum = "y" if self.ui.comp_sum_checkbox.isChecked() else "n"
        drilling_sum = "y" if self.ui.drill_sum_checkbox.isChecked() else "n"
        cement_log = "y" if self.ui.cement_log_checkbox.isChecked() else "n"
        logs = "y" if self.ui.logs_included_checkbox.isChecked() else "n"
        bhl_processed = "y"
        as_drilled_excel = "y" if self.ui.asdrilled_excel_checkbox.isChecked() else "n"
        edited_wcr = "y"
        all_edits = get_all_edits()

        # Prepare data row
        all_data = [days_average, date_filed.strftime("%m/%d/%Y"), ret_count, sundry_number,
                    api_number, well_name, date_processed.strftime("%m/%d/%Y"), company,
                    action_taken, comp_sum, drilling_sum, cement_log, logs, bhl_processed,
                    as_drilled_excel, edited_wcr, all_edits]

        # Write to spreadsheet
        letters = "abcdefghijklmnopqrstuvwxyz"
        for i in range(len(all_data)):
            sheet[letters[i] + str(max_row)] = all_data[i]

        wb.save(file)
        wb.close()
        self.savedPopup()

    def get_wcr_person_db_update(self) -> pd.DataFrame:
        """
        Retrieve WCR submission information for tracking updates.

        Returns:
            DataFrame containing the most recent WCR submission details

        This query gets the latest sundry submission for the well,
        which represents the most recent WCR filing.
        """
        query = f"""select wi.SundryNo, dsh.OperatorName, dsh.WellNameNumber, 
                   dsh.APINumber, dsh.LateralName as ConstructKey, tas.SubmitDate
        from tblAPDWCRWellInfo wi
        join [dbo].tblAPD ta on wi.APDNo = ta.APDNo
        join [dbo].[tblAPDSundry] tas on ta.API_WellNo = tas.APINO
        JOIN DirectionalSurveyHeader dsh ON LEFT(ta.API_WellNo, 10) = dsh.APINumber 
        where dsh.APINumber = {self.api} and dsh.LateralName = '{self.lateral}'
        order BY tas.SubmitDate"""

        wcr_info = self.db.query_to_dataframe(query)
        wcr_info = wcr_info.drop_duplicates(keep="first")
        wcr_info = wcr_info.tail(1)  # Get most recent submission
        return wcr_info

    def wcr_logs_sql(self) -> pd.DataFrame:
        """
        Retrieve well log submission records from the database.

        Returns:
            DataFrame containing log types and received dates

        Well logs are required attachments for WCR submissions,
        including cement bond logs, production logs, etc.
        """
        query = f"""select w.WellID, cl.LogType, cl.ReceivedDate
            from well w
            inner join construct c on w.pkey=c.WellKey
            inner join ConstructLog cl on cl.ConstructKey = c.PKey
            where w.WellID = '{self.api}'
            order by cl.ReceivedDate"""
        df = self.db.query_to_dataframe(query)
        return df

    def wcr_sundries_sql(self) -> pd.DataFrame:
        """
        Retrieve sundry submission records for the well.

        Returns:
            DataFrame containing submission types and dates

        Sundries include various regulatory filings like WCRs,
        workover reports, and other operational notices.
        """
        query = f"""select APINO, TypeSubmission, SubmitDate, Operations
                    from tblAPDSundry
        where APINO = '{self.api}0000'
        order by SubmitDate"""

        df = self.db.query_to_dataframe(query)
        # Format API number and dates for display
        df['APINO'] = df['APINO'].astype(str).apply(lambda row: row[:10])
        df['SubmitDate'] = df['SubmitDate'].astype(str).apply(lambda row: row[:10])
        return df

    def process_survey_data(self) -> None:
        """
        Process directional survey data to calculate coordinates at key well points.

        This method:
        1. Extracts Township-Range-Section data from concatenated strings
        2. Interpolates survey data at perforation depths
        3. Calculates footage from section lines
        4. Updates UI displays with processed data

        The processing ensures accurate coordinate reporting for regulatory compliance
        and proper well location documentation.
        """

        def transform_string(s: str) -> Dict[str, str]:
            """
            Parse concatenated location string into components.

            Args:
                s: Concatenated string in format "SSTTTRRRB"
                   where S=Section, T=Township, R=Range, B=Baseline

            Returns:
                Dictionary with parsed location components
            """
            section = str(int(s[:2]))
            township = str(int(s[2:4]))
            township_dir = s[4]
            rng = str(int(s[5:7]))
            rng_dir = s[7]
            baseline = s[-1]
            return {
                'Section': section,
                'Township': township,
                'Township_Direction': township_dir,
                'Range': rng,
                'Range_Direction': rng_dir,
                'Baseline': baseline
            }

        # Get survey data based on selected north reference
        wcr_df, kop = self.return_button_data()

        # Extract perforation interval depths from UI
        perf_mods = [float(self.ui.perf_interval_top.text()),
                     float(self.ui.perf_interval_bottom.text())]

        # Insert interpolated rows at perforation depths
        for i in perf_mods:
            wcr_df = self.insert_row_at_md(wcr_df, i)

        # Parse location strings into separate columns
        new_columns = wcr_df['Conc'].apply(transform_string).apply(pd.Series)
        wcr_df = wcr_df.join(new_columns)

        # Collect key depths for processing
        init = [kop['measured_depth'].iloc[0]]
        self.perf_mods = perf_mods
        used_depths = perf_mods + init
        perf_mods = wcr_df[wcr_df['measured_depth'].isin(used_depths)]

        # Process survey data and update displays
        wcr_result_df = self.process_survey_data_misc(kop, perf_mods, wcr_df)
        self.wcr_result_df = wcr_result_df
        self.data_writer_process(wcr_result_df)
        self.set_shl_bhl_data(wcr_df)
        self.write_coords(wcr_result_df)

    def insert_row_at_md(self, df: pd.DataFrame, target_md: float) -> pd.DataFrame:
        """
        Insert an interpolated row at a specific measured depth in the survey data.

        Args:
            df: DataFrame containing survey data
            target_md: Target measured depth for interpolation

        Returns:
            DataFrame with interpolated row inserted

        This method performs linear interpolation for numeric columns and
        carries forward string values from the preceding row. It's critical
        for calculating coordinates at perforation depths that may not have
        been surveyed directly.
        """
        # Ensure dataframe is sorted by measured depth
        df = df.sort_values('measured_depth').reset_index(drop=True)

        # Find insertion point
        insert_index = df['measured_depth'].searchsorted(target_md)

        # Handle edge cases
        if insert_index == 0:
            return df  # Target is above survey start

        if insert_index == len(df):
            # Extrapolate using last two rows
            lower_row = df.iloc[-2]
            upper_row = df.iloc[-1]
        else:
            # Interpolate between surrounding rows
            lower_row = df.iloc[insert_index - 1]
            upper_row = df.iloc[insert_index]

        # Initialize new row
        new_row = {'measured_depth': target_md}

        # Process each column
        for col in df.columns:
            if col == 'measured_depth':
                continue

            # Handle string columns
            if df[col].dtype == 'object':
                new_row[col] = lower_row[col]
            else:
                try:
                    # Linear interpolation for numeric columns
                    new_row[col] = np.interp(
                        target_md,
                        [lower_row['measured_depth'], upper_row['measured_depth']],
                        [lower_row[col], upper_row[col]]
                    )
                except (TypeError, ValueError):
                    # Fall back to lower row value if interpolation fails
                    new_row[col] = lower_row[col]

        # Insert the new row
        new_row = pd.Series(new_row)
        df = pd.concat([df.iloc[:insert_index],
                        pd.DataFrame([new_row]),
                        df.iloc[insert_index:]]).reset_index(drop=True)

        return df

    def data_writer_process(self, wcr_result_df: pd.DataFrame) -> None:
        """
        Update UI elements with processed survey data.

        Args:
            wcr_result_df: DataFrame containing processed survey results

        This method populates the UI with coordinate data and sets
        the appropriate north reference display.
        """
        # Map single-letter codes to full baseline names
        baseline_dict = {'T': 'True', 'G': 'Grid'}

        # Define UI element groups for as-drilled survey data
        asdrilled_survey_data = [
            [self.ui.bhl1_depth, self.ui.bhl1_ns_offset, self.ui.bhl1_ns_offset_combo,
             self.ui.bhl1_ew_offset, self.ui.bhl1_ew_offset_combo],
            [self.ui.bhl1_depth_2, self.ui.bhl1_ns_offset_2, self.ui.bhl1_ns_offset_combo_2,
             self.ui.bhl1_ew_offset_2, self.ui.bhl1_ew_offset_combo_2],
            [self.ui.bhl2_depth, self.ui.bhl2_ns_offset, self.ui.bhl2_ns_offset_combo,
             self.ui.bhl2_ew_offset, self.ui.bhl2_ew_offset_combo],
            [self.ui.bhl2_depth_2, self.ui.bhl2_ns_offset_2, self.ui.bhl2_ns_offset_combo_2,
             self.ui.bhl2_ew_offset_2, self.ui.bhl2_ew_offset_combo_2]
        ]

        # Set north reference combo box
        one_letter_id = self.north_ref[0].upper()
        self.ui.survey_north_ref_combo.setCurrentText(baseline_dict[one_letter_id])

    def process_survey_data_misc(
            self,
            kop: pd.DataFrame,
            perf_mods: pd.DataFrame,
            wcr_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Combine survey data from different well points into final result.

        Args:
            kop: DataFrame containing Kick-Off Point data
            perf_mods: DataFrame containing perforation interval data
            wcr_df: Complete survey DataFrame

        Returns:
            Combined DataFrame with SHL, perforation points, and BHL

        This creates the final survey dataset with all required points
        for WCR reporting.
        """
        # Extract surface and bottom hole locations
        last_line = wcr_df.tail(1)
        first_line = wcr_df.head(1)
        first_and_last_df = pd.concat([first_line, last_line], ignore_index=True)

        # Combine all key points and sort by measured depth
        wcr_result_df = pd.concat([perf_mods, first_and_last_df], ignore_index=True).sort_values(
            'measured_depth').reset_index(drop=True)

        return wcr_result_df

    def return_button_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Get survey data based on selected north reference type.

        Returns:
            Tuple of (survey DataFrame, KOP DataFrame)

        This method returns either True North or Grid North referenced
        survey data based on user selection.
        """
        # Get selected north reference from radio buttons
        active_button_id = self.ui.wcr_survey_northref_buttons.checkedId()
        used_spec_survey = self.spec_surveys['drl_df']

        # Return appropriate survey data
        if active_button_id == -2:  # True North
            wcr_df = self.df['drl_df_true_dx'].clearance_data
            kop = used_spec_survey[
                (used_spec_survey['Point'] == 'KOP') &
                (used_spec_survey['type'] == 'drl_df_true_dx')
                ]
        else:  # Grid North
            wcr_df = self.df['drl_df_grid_dx'].clearance_data
            kop = used_spec_survey[
                (used_spec_survey['Point'] == 'KOP') &
                (used_spec_survey['type'] == 'drl_df_grid_dx')
                ]
        return wcr_df, kop

    def set_shl_bhl_data(self, wcr_df: pd.DataFrame) -> None:
        """
        Calculate and display latitude/longitude for surface and bottom hole locations.

        Args:
            wcr_df: DataFrame containing survey data

        Converts UTM coordinates to lat/lon for regulatory reporting
        and updates the UI display fields.
        """
        # Get surface and bottom hole rows
        shl_row = wcr_df.head(1)
        bhl_row = wcr_df.tail(1)

        # Convert UTM to lat/lon (Zone 12T for Utah)
        latlon_shl = utm.to_latlon(shl_row['easting'], shl_row['northing'], 12, 'T')
        latlon_bhl = utm.to_latlon(bhl_row['easting'], bhl_row['northing'], 12, 'T')

        # Update UI with rounded coordinates
        self.ui.shl_lat.setText(str(abs(round(float(latlon_shl[0].iloc[0]), 5))))
        self.ui.shl_lon.setText(str(abs(round(float(latlon_shl[1].iloc[0]), 5))))
        self.ui.bhl_lat_2.setText(str(abs(round(float(latlon_bhl[0].iloc[0]), 5))))
        self.ui.bhl_lon_2.setText(str(abs(round(float(latlon_bhl[1].iloc[0]), 5))))

    def write_coords(self, df: pd.DataFrame) -> None:
        """
        Display coordinate data in the UI table widget.

        Args:
            df: DataFrame containing coordinate data to display

        Populates the coordinate table with UTM coordinates and
        footage from section lines for all key well points.
        """
        # Clear existing data
        self.coords_wcr_model.setRowCount(0)
        self.ui.display_table_utm_locs.setModel(self.coords_wcr_model)
        self.ui.display_table_utm_locs.setUpdatesEnabled(False)

        # Select display columns
        used_dataframe = df[['measured_depth', 'tvd', 'easting', 'northing',
                             'FNL', 'FSL', 'FEL', 'FWL']]
        columns = ['measured_depth', 'tvd', 'easting', 'northing',
                   'FNL', 'FSL', 'FEL', 'FWL']
        used_dataframe = used_dataframe.drop_duplicates(keep="first")

        # Populate table model
        for i, row in used_dataframe.iterrows():
            # Round values for display
            row_data = [str(int(round(r, 0))) for r in row]
            items = [QStandardItem(item) for item in row_data]

            # Store actual values as data
            for column_index, value in enumerate(row):
                items[column_index].setData(int(round(value, 0)))
            self.coords_wcr_model.appendRow(items)

        # Set headers and configure display
        self.coords_wcr_model.setHorizontalHeaderLabels(columns)
        self.ui.display_table_utm_locs.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.display_table_utm_locs.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.display_table_utm_locs.verticalHeader().setVisible(False)
        self.ui.display_table_utm_locs.setShowGrid(True)

        # Re-enable updates
        self.ui.display_table_utm_locs.setUpdatesEnabled(True)
        self.ui.display_table_utm_locs.show()

    def wcr_perf_mods(self) -> Tuple[List[float], str]:
        """
        Retrieve perforation interval data from the database.

        Returns:
            Tuple of (perforation depths list, perforation date string)

        Gets the primary perforation zone for the well, which is
        required information for WCR reporting.
        """
        try:
            sql_query = f"""select [Top], Bottom, [Perf Date]
                        from well w 
                        join construct c on w.PKey = c.WellKey  
                        join [dbo].[vwDM_ConstructPerf] cp on c.PKey = cp.PKey
                        where wellid = {self.api} and [Zone Type] = 'Perforations'"""

            tops = self.db.query_to_dataframe(sql_query)
            tops = tops.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

            # Extract perforation data
            perf_top = tops['Top'].values.tolist()[0]
            perf_bottom = tops['Bottom'].values.tolist()[0]
            perf_mods = [float(perf_top), float(perf_bottom)]
            perf_date = tops['Perf Date'].iloc[0]
        except IndexError:
            # Default values if no perforations found
            perf_mods = [0, 0]
            perf_date = "1/1/1900"

        return perf_mods, perf_date

    def process_wcr(self) -> None:
        """
        Initialize WCR processing by setting up UI elements.

        This method is called to start the WCR workflow and
        prepares all combo boxes with appropriate options.
        """
        self.setup_combo_boxes()

    def present_plat_data(self) -> None:
        """
        Display plat (planned location) data in the UI.

        Populates all location fields with data from the well's
        Application for Permit to Drill (APD), including:
        - Township, Range, Section coordinates
        - Footage from section lines
        - Meridian and baseline information
        """
        wcr_tsr_db = self.loc_df
        self.ui.display_data.clear()

        # Extract location data by zone
        sd = wcr_tsr_db[(wcr_tsr_db['zone_name'] == 'Surface Location')]
        td = wcr_tsr_db[(wcr_tsr_db['zone_name'] == 'Proposed Depth')]
        prod = wcr_tsr_db[(wcr_tsr_db['zone_name'] == 'Uppermost Producing')]

        baseline_dict = {'U': 'Uintah', 'S': 'Salt Lake'}

        # Set footage from north/south lines
        planned_survey_data_fnsl = [
            [self.ui.shl_fnsl_data.setText(str(sd.iloc[0, 6])),
             self.ui.shl_fnsl_combo.setCurrentText(str(sd.iloc[0, 7]))],
            [self.ui.bhl1_fnsl_data.setText(str(prod.iloc[0, 6])),
             self.ui.bhl1_fnsl_combo.setCurrentText(str(prod.iloc[0, 7]))],
            [self.ui.bhl2_fnsl_data.setText(str(td.iloc[0, 6])),
             self.ui.bhl2_fnsl_combo.setCurrentText(str(td.iloc[0, 7]))]
        ]

        # Set footage from east/west lines
        planned_survey_data_fewl = [
            [self.ui.shl_fewl_data.setText(str(sd.iloc[0, 8])),
             self.ui.shl_fewl_combo.setCurrentText(str(sd.iloc[0, 9]))],
            [self.ui.bhl1_fewl_data.setText(str(prod.iloc[0, 8])),
             self.ui.bhl1_fewl_combo.setCurrentText(str(prod.iloc[0, 9]))],
            [self.ui.bhl2_fewl_data.setText(str(td.iloc[0, 8])),
             self.ui.bhl2_fewl_combo.setCurrentText(str(td.iloc[0, 9]))]
        ]

        # Set Township-Range-Section data
        tsr_data = [
            [self.ui.shl_section.setText(str(sd.iloc[0, 0])),
             self.ui.shl_township.setText(str(sd.iloc[0, 1])),
             self.ui.shl_township_dir_combo.setCurrentText(str(sd.iloc[0, 2])),
             self.ui.shl_range.setText(str(sd.iloc[0, 3])),
             self.ui.shl_range_dir_combo.setCurrentText(str(sd.iloc[0, 4]))],
            [self.ui.bhl1_section.setText(str(prod.iloc[0, 0])),
             self.ui.bhl1_township.setText(str(prod.iloc[0, 1])),
             self.ui.bhl1_township_dir_combo.setCurrentText(str(prod.iloc[0, 2])),
             self.ui.bhl1_range.setText(str(prod.iloc[0, 3])),
             self.ui.bhl1_range_dir_combo.setCurrentText(str(prod.iloc[0, 4]))],
            [self.ui.bhl2_section.setText(str(td.iloc[0, 0])),
             self.ui.bhl2_township.setText(str(td.iloc[0, 1])),
             self.ui.bhl2_township_dir_combo.setCurrentText(str(td.iloc[0, 2])),
             self.ui.bhl2_range.setText(str(td.iloc[0, 3])),
             self.ui.bhl2_range_dir_combo.setCurrentText(str(td.iloc[0, 4]))]
        ]

        # Set meridian baseline
        self.ui.meridian_combo.setCurrentText(baseline_dict[sd['baseline'].values[0]])

    def create_well_logs_list(self) -> None:
        """
        Populate the logs report table with well log submission data.

        Displays all logs submitted for the well, including:
        - Log type (cement bond, production, etc.)
        - Received date

        This helps verify all required logs are included with the WCR.
        """
        # Clear and prepare table
        self.report_box_model.setRowCount(0)
        self.ui.logs_report_box.setModel(self.report_box_model)
        self.ui.logs_report_box.setUpdatesEnabled(False)

        # Get log data
        df = self.wcr_logs_sql()
        columns = df.columns

        if df.empty:
            # Show message if no logs found
            self.report_box_model.setColumnCount(1)
            empty_item = QStandardItem("No logs found")
            self.report_box_model.appendRow([empty_item])
        else:
            # Populate table with log data
            for i, row in df.iterrows():
                items = [QStandardItem(str(item)) for item in row]
                for column_index, value in enumerate(row):
                    items[column_index].setData(value)
                self.report_box_model.appendRow(items)
            self.report_box_model.setHorizontalHeaderLabels(columns)

        # Configure table display
        self.ui.logs_report_box.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.logs_report_box.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.logs_report_box.verticalHeader().setVisible(False)
        self.ui.logs_report_box.setShowGrid(True)

        # Re-enable updates
        self.ui.logs_report_box.setUpdatesEnabled(True)
        self.ui.logs_report_box.show()

    def create_sundries_list(self) -> None:
        """
        Populate sundries combo box with available sundry submissions.

        Sundries are miscellaneous regulatory filings that may contain
        relevant information for WCR preparation.
        """
        df = self.wcr_sundries_sql()
        self.ui.sundries_combo_box.clear()
        self._set_sundries_df(df)

        # Add each sundry to combo box with date and type
        for idx, row in df.iterrows():
            label = f"""{row['SubmitDate']} {row['TypeSubmission']}"""
            self.ui.sundries_combo_box.addItem(label)

    def _set_sundries_df(self, value: pd.DataFrame) -> None:
        """
        Store sundries DataFrame for internal use.

        Args:
            value: DataFrame containing sundries data
        """
        self._sundries_df = value

    def sundries_combo_box_process(self) -> None:
        """
        Display selected sundry details in the report box.

        Shows the operations text from the selected sundry submission,
        which may contain relevant completion or workover information.
        """
        self.ui.sundries_report_box.clear()
        df_index = self._sundries_df.iloc[self.ui.sundries_combo_box.currentIndex()]['Operations']
        self.ui.sundries_report_box.setText(df_index)

    def setup_combo_boxes(self) -> None:
        """
        Initialize all combo boxes with appropriate option lists.

        Sets up dropdown menus for:
        - North references (True/Grid)
        - Directions (N/S/E/W)
        - Section line references (FNL/FSL/FEL/FWL)
        - Meridians and baselines

        This ensures consistent data entry options throughout the UI.
        """
        # Define all dropdown options
        combo_options = {
            'north_ref': ['True', 'Grid'],
            'fnsl': ['FNL', 'FSL'],
            'fewl': ['FEL', 'FWL'],
            'ns': ['N', 'S'],
            'ew': ['E', 'W'],
            'meridian': ['Salt Lake', 'Uintah'],
            'directions': ['SE', 'NE', 'SW', 'NW']
        }

        # Group related combo boxes by function
        widget_groups = {
            'north_ref': [
                self.ui.plat_north_ref_combo,
                self.ui.survey_north_ref_combo
            ],
            'fnsl': [
                self.ui.shl_fnsl_combo,
                self.ui.bhl1_fnsl_combo,
                self.ui.bhl2_fnsl_combo
            ],
            'fewl': [
                self.ui.shl_fewl_combo,
                self.ui.bhl1_fewl_combo,
                self.ui.bhl2_fewl_combo
            ],
            'township_ns': [
                self.ui.shl_township_dir_combo,
                self.ui.bhl1_township_dir_combo,
                self.ui.bhl2_township_dir_combo
            ],
            'range_ew': [
                self.ui.shl_range_dir_combo,
                self.ui.bhl1_range_dir_combo,
                self.ui.bhl2_range_dir_combo
            ],
            'meridian': [
                self.ui.meridian_combo
            ],
            'ns_offset': [
                self.ui.bhl1_ns_offset_combo,
                self.ui.bhl1_ns_offset_combo_2,
                self.ui.bhl2_ns_offset_combo,
                self.ui.bhl2_ns_offset_combo_2
            ],
            'ew_offset': [
                self.ui.bhl1_ew_offset_combo,
                self.ui.bhl1_ew_offset_combo_2,
                self.ui.bhl2_ew_offset_combo,
                self.ui.bhl2_ew_offset_combo_2
            ],
            'directions': [
                self.ui.shl_to_bhl1_dir_combo,
                self.ui.bhl1_to_bhl2__dir_combo,
                self.ui.bhl_1_bearing_dir_ew_combo,
                self.ui.bhl_1_bearing_dir_ns_combo,
                self.ui.bhl_2_bearing_dir_ew_combo,
                self.ui.bhl_2_bearing_dir_ns_combo
            ]
        }

        # Populate all combo boxes with appropriate options
        for group_name, widgets in widget_groups.items():
            # Determine which options to use
            options = combo_options.get(group_name, [])
            if group_name in ['township_ns']:
                options = combo_options['ns']
            elif group_name in ['range_ew', 'ew_offset']:
                options = combo_options['ew']
            elif group_name in ['ns_offset']:
                options = combo_options['ns']

            # Add options to each widget in the group
            for widget in widgets:
                widget.addItems(options)

    def interpolation_table_process(self) -> None:
        """
        Calculate interpolated coordinates between two survey stations.

        This method performs linear interpolation to find coordinates
        at a specific depth between two known survey points. It's used
        when exact survey data isn't available at required depths.

        The interpolation assumes a straight line between survey stations,
        which is standard practice for short intervals in directional drilling.
        """
        try:
            # Get survey station data from UI
            next_depth = float(self.ui.next_depth_line.text())
            next_ns = float(self.ui.next_ns_line.text())
            next_ew = float(self.ui.next_ew_line.text())
            prev_depth = float(self.ui.prev_depth_line.text())
            prev_ns = float(self.ui.prev_ns_line.text())
            prev_ew = float(self.ui.prev_ew_line.text())
            target_depth = float(self.ui.cur_depth_line.text())

            # Linear interpolation formula: y = y1 + (x - x1) * (y2 - y1) / (x2 - x1)
            interpolated_ns = ((next_ns - prev_ns) / (next_depth - prev_depth)) * (target_depth - prev_depth) + prev_ns
            interpolated_ew = ((next_ew - prev_ew) / (next_depth - prev_depth)) * (target_depth - prev_depth) + prev_ew

            # Update UI with calculated values
            self.ui.cur_ns_line.setText((str(round(interpolated_ns, 2))))
            self.ui.cur_ew_line.setText((str(round(interpolated_ew, 2))))
        except ValueError:
            # Silently handle invalid input during typing
            pass