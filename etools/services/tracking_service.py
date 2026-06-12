"""Personal WCR tracking workbook (TrackingWCR.xlsx).

Replicates the V2 "Update Personal Record" feature: one row per processed
WCR keyed by API (column E), recording when it was filed, when you
processed it, what the submission included, and what you had to edit.
Column layout matches the V2-era workbook (A:Q) so an existing file keeps
working — copy your old TrackingWCR.xlsx next to the app to keep history.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl

from etools.logging_setup import get_logger

log = get_logger(__name__)

_HEADERS = (
    "Days, average no returns",
    "Date Filed",
    "RET? How many times?",
    "Sundry #",
    "API",
    "Well Name",
    "Date Processed",
    "Company",
    "Action Taken",
    "Comp Sum",
    "Drilling Sum",
    "CMT",
    "Logs",
    "BHL",
    "As drilled Excel Survey",
    "Edited WCR",
    "Edits",
)


def update_tracking_workbook(
    *,
    path: str | Path,
    api: str,
    well_name: str | None,
    operator: str | None,
    sundry_no: str | int | None,
    date_filed: datetime | None,
    returns: int = 0,
    action_taken: bool = False,
    comp_sum: bool = False,
    drill_sum: bool = False,
    cement_log: bool = False,
    logs_included: bool = False,
    as_drilled_excel: bool = False,
    edits: list[str] | None = None,
) -> Path:
    """Insert or update the row for ``api``. Returns the workbook path.

    Raises PermissionError when the workbook is open in Excel.
    """
    path = Path(path)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "TrackingWCR"
        for col, header in enumerate(_HEADERS, start=1):
            sheet.cell(row=1, column=col, value=header)

    api_int = int(str(api)[:10])
    api_cells = [c.value for c in sheet["E"]]
    if api_int in api_cells:
        target_row = api_cells.index(api_int) + 1
    else:
        target_row = sheet.max_row + 1

    today = date.today()
    days_average = None
    if date_filed is not None:
        days_average = int(abs(date_filed.date() - today).days)

    def yn(flag: bool) -> str:
        return "y" if flag else "n"

    values = [
        days_average,
        date_filed.strftime("%m/%d/%Y") if date_filed else None,
        int(returns or 0),
        sundry_no,
        api_int,
        well_name,
        today.strftime("%m/%d/%Y"),
        operator,
        yn(action_taken),
        yn(comp_sum),
        yn(drill_sum),
        yn(cement_log),
        yn(logs_included),
        "y",  # BHL processed — always true when generating from this app
        yn(as_drilled_excel),
        "y" if (edits or action_taken) else "n",
        "/".join(edits or []),
    ]
    for col, value in enumerate(values, start=1):
        sheet.cell(row=target_row, column=col, value=value)

    wb.save(path)
    wb.close()
    log.info("tracking.updated", path=str(path), api=api_int, row=target_row)
    return path
