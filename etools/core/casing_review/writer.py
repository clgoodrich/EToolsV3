"""Write a fully-computed ``CasingDesign`` into the Casing Review xlsx.

Drops computed values into the Casing Review sheet AND the DataPrint
panel, so the workbook opens with every design factor already filled
in (the formulas still recompute when Excel opens the file — we just
don't depend on them).
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

from etools.core.casing_review.bope import build_bope_review
from etools.core.casing_review.domain import CasingDesign, CasingStringDesign
from etools.core.casing_review.footages import (
    footages_to_xy,
    location_footages,
    polygon_footages,
)
from etools.core.casing_review.generator import CASING_REVIEW_TEMPLATE
from etools.core.casing_review.grid_corners import derive_section_corners
from etools.core.casing_review.sections import PLSSKey
from etools.logging_setup import get_logger

log = get_logger(__name__)


# Row offsets relative to each STRING block top (10, 25, 40, 55).
_DATA_ROW_OFFSET = 2  # row 12 / 27 / 42 / 57


def write_casing_review(
    design: CasingDesign,
    output_path: Path,
    *,
    template_path: Path | None = None,
    surface_location=None,
    producing_interval_location=None,
    td_location=None,
    intermediate_locations: list | None = None,
    section_locations: list | None = None,
    dx_survey_locations: list | None = None,
    dx_survey_footages: list | None = None,
    plat_repo=None,
    bope_system_psi: float | None = None,
) -> Path:
    """Fill the Casing Review xlsx with both inputs and computed values.

    Section-sheet inputs:
        * ``surface_location``              → SHL Section (PLSS + UTM block)
        * ``producing_interval_location``   → BHL Section 1
        * ``td_location``                   → BHL Section 3
                                              (BHL 2 left blank for an
                                              intermediate-section pass,
                                              wired in when clearance data
                                              is plumbed through)

    All three are ``APDLocationRow`` instances from
    ``APDPdfData.locations`` (Section 20 of the Form 3).
    """
    template_path = Path(template_path or CASING_REVIEW_TEMPLATE)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    cr = wb["Casing Review"]

    # Header
    cr["B4"] = design.company
    cr["B5"] = design.well_name
    cr["B6"] = design.api
    cr["B9"] = design.frac_gradient_psi_per_ft

    block_tops = (10, 25, 40, 55)
    for idx, s in enumerate(design.strings[:4]):
        _write_block(cr, block_tops[idx], s, is_surface=idx == 0)

    # Footer block per string (Next Set Depth / Next MW / Next BHP / Fracture
    # Initiation Pressure / Max anticipated pressure at shoe). The template
    # leaves these as formulas, but openpyxl writes formulas with NO cached
    # value, so any viewer that doesn't recalc on open shows them blank.
    # Pre-fill them as values (like the rest of the computed cells) so they
    # always display, and blank the footers of any unused string slot.
    _write_footer_values(cr, design)

    # BOPE sheet: fill its three hand-entered inputs (previous-shoe depth,

    # BOPE sheet: fill its three hand-entered inputs (previous-shoe depth,
    # proposed BOPE rating per string, operator's max anticipated pressure).
    # The rest of the sheet is formula-driven off Casing Review and now
    # resolves once the per-string TVDs are written above.
    if "BOPE" in wb.sheetnames:
        _write_bope(wb["BOPE"], design, bope_system_psi=bope_system_psi)

    # DataPrint panel mirrors per-string outputs into a normalized form.
    if "DataPrint" in wb.sheetnames:
        _write_dataprint(wb["DataPrint"], design)

    # SHL + BHL Section sheets. Each section sheet's row-7 input block
    # drives every formula in that sheet via Grid-Numbers DGET lookups.
    if section_locations:
        # The DGET lookups resolve only for sections present in the
        # embedded "Grid Numbers" sheet (a curated subset). Backfill any
        # crossed section missing from it by deriving its 16 quarter-side
        # rows from the plat polygon — otherwise the sheet's bearings come
        # up blank for every section the well merely passes through.
        if plat_repo is not None and "Grid Numbers" in wb.sheetnames:
            _ensure_grid_numbers_coverage(
                wb["Grid Numbers"], section_locations, plat_repo
            )
        # The BHL Section sheets auto-detect which sections the wellbore
        # crosses by walking the survey path stored in DxSurvey rows 8-10
        # (K.O. Point / Prod. Interval / Total Depth). Populate those
        # offsets so the native detection resolves — without them every
        # BHL sheet's bearing grid comes up blank or #VALUE!.
        if dx_survey_locations and "DxSurvey" in wb.sheetnames:
            _write_dx_survey_locations(wb["DxSurvey"], dx_survey_locations)
        _write_section_sheets_from_traversal(
            wb, design, section_locations, plat_repo=plat_repo
        )
        # The native survey-path walk that fills the KOP/Landing/Total-Depth
        # "Section Line Footages" (I/K columns) is unreliable and blanks out
        # the FINAL (TD) footages on cross-township/excursion wells. Write
        # them directly from the clearance data so the summary always shows
        # the bottom-hole footages.
        if dx_survey_footages:
            _write_path_footages(wb, section_locations, dx_survey_footages)
    else:
        _write_section_sheets_legacy(
            wb,
            design,
            surface_location=surface_location,
            producing_interval_location=producing_interval_location,
            td_location=td_location,
            intermediate_locations=intermediate_locations,
            plat_repo=plat_repo,
        )

    wb.save(output_path)
    return output_path


# Section sheets, in order. Index 0 is the surface section; the rest are
# the bottom-hole crossings. The template ships SHL + BHL 1-3; we ensure
# BHL 1-7 always exist (8 sheets) so any wellbore's full section list fits.
_TEMPLATE_SECTION_SHEETS = (
    "SHL Section",
    "BHL Section 1",
    "BHL Section 2",
    "BHL Section 3",
)
_MAX_BHL_SHEETS = 7  # SHL + BHL 1-7 = 8 section slots


def _section_sheet_name(idx: int) -> str:
    """0 -> 'SHL Section'; N -> 'BHL Section N'."""
    return "SHL Section" if idx == 0 else f"BHL Section {idx}"


def _all_section_sheet_names() -> list[str]:
    return ["SHL Section"] + [f"BHL Section {i}" for i in range(1, _MAX_BHL_SHEETS + 1)]


def _split_top_args(s: str) -> list[str]:
    """Split a formula-body on top-level commas, respecting parens/quotes."""
    args: list[str] = []
    depth = 0
    in_q = False
    cur = ""
    for ch in s:
        if ch == "'":
            in_q = not in_q
        if not in_q:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                args.append(cur)
                cur = ""
                continue
        cur += ch
    if cur:
        args.append(cur)
    return args


def _peel_continuation(formula: str) -> str:
    """Reduce a cross-section *continuation* formula to its own-section value.

    The Casing Review section sheets wrap each bearing cell in up to three
    ``IF(AND($L$38='SHL Section'!$BE$nn, MAX(<prev sheet>!$BF$..)=k), <prev
    sheet value>, …)`` layers that try to inherit the matching boundary
    bearing from the *previous* crossed section. That cross-sheet walk
    breaks when the previous section sits in a different township (its
    ``$BE$``/``$BF$`` helpers resolve to ``#VALUE!``/``#N/A``), and the
    error lands in the visible grid. Peeling those adjacency layers off
    leaves the cell's own-section fallback (its own DGET / own computation),
    which is exactly the bearing for *this* section.
    """
    f = formula[1:] if formula.startswith("=") else formula
    while f.startswith("IF(") and f.endswith(")"):
        args = _split_top_args(f[3:-1])
        if len(args) != 3 or "$BE$" not in args[0]:
            break
        f = args[2].strip()
    return "=" + f


def _disable_continuation(ws) -> int:
    """Strip cross-section continuation from every bearing cell on a sheet.

    Used for wells that thread a section in a different township from the
    surface, where the template's continuation walk would otherwise leave
    ``#VALUE!`` in the visible grid. Each affected section then displays its
    own bearings. Returns the number of cells rewritten.
    """
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "'SHL Section'!$BE$" in v:
                cell.value = _peel_continuation(v)
                n += 1
    return n


# The visible bearing grid the user reads (and what the sheet's print area
# captures). Unused section sheets are blanked over this range so they stay
# present/visible but don't show a #VALUE! grid.
_GRID_MIN_ROW = 15
_GRID_MAX_ROW = 63
_GRID_MAX_COL = 13  # column M


def _blank_unused_section_sheet(ws) -> int:
    """Clear the bearing-grid formulas on a section sheet the bore never uses.

    An unused BHL sheet has no real section assigned (its ``L38``/``N7`` stay
    at the copied template default), so every DGET-driven bearing cell
    resolves to ``#VALUE!``. The user wants all eight sheets present and
    visible (never hidden), but a screen of ``#VALUE!`` is noise — so we drop
    the formula cells across the visible grid, leaving a clean, empty sheet.
    The well/API header (rows 2-3) and the four reference-point rows (7-10,
    which resolve fine off DxSurvey/SHL) are left intact. Returns the number
    of cells cleared.
    """
    n = 0
    for row in ws.iter_rows(
        min_row=_GRID_MIN_ROW, max_row=_GRID_MAX_ROW, max_col=_GRID_MAX_COL
    ):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None
                n += 1
    return n


def _is_cross_township(loc, shl) -> bool:
    """True if ``loc`` sits in a different township/range/meridian than SHL."""
    def key(x):
        return (
            str(getattr(x, "township", "") or "").upper(),
            str(getattr(x, "township_dir", "") or "").upper(),
            str(getattr(x, "range", "") or "").upper(),
            str(getattr(x, "range_dir", "") or "").upper(),
            str(getattr(x, "meridian", "") or "").upper(),
        )
    return key(loc) != key(shl)


def _ensure_section_sheets(wb) -> None:
    """Make sure SHL + BHL 1-7 all exist, copying BHL Section 3 as needed.

    The template only ships BHL 1-3; we provision up to BHL 7 so a wellbore
    that threads many sections always has a sheet per section. Copies keep
    BHL Section 3's full formula structure; the section each one shows is
    driven by its own ``L38`` (set per-section by the writer), not by the
    copied adjacency references.
    """
    if "BHL Section 3" not in wb.sheetnames:
        return
    for idx in range(4, _MAX_BHL_SHEETS + 1):
        name = f"BHL Section {idx}"
        if name not in wb.sheetnames:
            new_sheet = wb.copy_worksheet(wb["BHL Section 3"])
            new_sheet.title = name
            log.info("section_sheet.created", name=name)


def _write_section_sheets_from_traversal(
    wb, design: CasingDesign, section_locations: list, *, plat_repo=None
) -> None:
    """Fill one section sheet per UNIQUE section the wellbore passes through.

    The wellbore's section traversal (:func:`build_section_traversal`) lists
    every distinct PLSS section in first-entry order — surface first, then
    each new section as the bore threads them (a detour into a neighbouring
    township-line section gets its own entry; only re-entries of an
    already-seen section are skipped). We write each one to the next
    sequential sheet (SHL, BHL 1, BHL 2, …), filling its full PLSS identity
    (section / township / range / meridian) so the bearing-grid DGET
    resolves that section — even when it sits in a different township from
    the surface.

    This deliberately does NOT rely on the template's built-in section
    auto-detection, which is unreliable (it misses the deepest section and
    can invent neighbouring sections the bore never enters). The
    ``#VALUE!`` that the native adjacency walk leaves in off-screen helper
    columns doesn't reach the visible bearing grid, which is driven purely
    by the per-section DGET.

    All eight section sheets (SHL + BHL 1-7) stay present and visible
    regardless of how many the wellbore actually uses — unused sheets are
    left in their template state, not hidden.

    ``section_locations[i]`` is the ``APDLocationRow`` for the i-th crossed
    section (0 = surface).
    """
    _ensure_section_sheets(wb)
    all_names = _all_section_sheet_names()

    crossed = list(section_locations or [])
    if len(crossed) > len(all_names):
        log.warning(
            "section_sheet.exceeds_capacity",
            crossed=len(crossed),
            capacity=len(all_names),
        )
        crossed = crossed[: len(all_names)]

    # Header on every section sheet; keep them all visible.
    for name in all_names:
        if name in wb.sheetnames:
            ws = wb[name]
            ws["C2"] = design.well_name
            ws["C3"] = design.api
            ws.sheet_state = "visible"

    # Write each crossed section's full PLSS into its sequential sheet.
    for idx, loc in enumerate(crossed):
        name = _section_sheet_name(idx)
        if name in wb.sheetnames:
            _write_section_sheet(
                wb[name], design, loc, sheet_label=name, plat_repo=plat_repo
            )

    # Sheets beyond the crossed count have no real section — blank their
    # bearing grids so they stay visible but clean (no #VALUE! noise).
    blanked = 0
    for idx in range(len(crossed), len(all_names)):
        name = _section_sheet_name(idx)
        if name in wb.sheetnames:
            blanked += _blank_unused_section_sheet(wb[name])
    if blanked:
        log.info(
            "section_sheet.unused_blanked",
            sheets=len(all_names) - len(crossed),
            cells=blanked,
        )

    # If the wellbore threads any section in a different township than the
    # surface, the template's cross-section continuation walk can't follow
    # it and leaves #VALUE! in the visible grid of the *following* sheet.
    # Disable the continuation on every BHL sheet so each section shows its
    # own bearings. Wells that stay in one township keep the native
    # continuation untouched.
    shl = crossed[0] if crossed else None
    if shl is not None and any(_is_cross_township(loc, shl) for loc in crossed[1:]):
        peeled = 0
        for name in all_names[1:]:  # BHL sheets only
            if name in wb.sheetnames:
                peeled += _disable_continuation(wb[name])
        log.info("section_sheet.continuation_disabled", cells=peeled)


def _write_dx_survey_locations(dxs_ws, dx_survey_locations: list) -> None:
    """Write the K.O./Prod-Interval/Total-Depth path offsets into DxSurvey.

    ``dx_survey_locations`` is up to three ``(md, n_offset, e_offset)``
    tuples (feet; N positive / S negative, E positive / W negative) written
    to rows 8, 9, 10 — columns C (MD), D (N/S), E (E/W). These are the
    inputs the section sheets' path-detection walks
    (``E8=ABS(DxSurvey!D8)``), so they must be present for the BHL sheets
    to resolve which sections the wellbore crosses.
    """
    for row, loc in zip((8, 9, 10), dx_survey_locations):
        if loc is None:
            continue
        md, n_off, e_off = loc
        if md is not None:
            dxs_ws.cell(row, 3, round(float(md), 2))
        if n_off is not None:
            dxs_ws.cell(row, 4, round(float(n_off), 4))
        if e_off is not None:
            dxs_ws.cell(row, 5, round(float(e_off), 4))


def _write_path_footages(wb, section_locations, dx_survey_footages) -> None:
    """Write KOP/Landing/TD section-line footages straight into the sheets.

    ``dx_survey_footages`` is ``[(conc, {fnl,fsl,fel,fwl}) | None]`` for the
    KOP / Landing / Total-Depth stations (rows 8 / 9 / 10). Each footage is
    relative to the section that station sits in, so it's written to:
      * the ``SHL Section`` summary (which natively aggregates these), and
      * the section sheet whose section matches the station's ``conc``
        (its own row, where the template would otherwise show it).
    Convention matches the surface row: I = N/S footage, J = 1(FNL)/2(FSL);
    K = E/W footage, L = 1(FEL)/2(FWL).
    """
    if "SHL Section" not in wb.sheetnames:
        return
    shl = wb["SHL Section"]

    # Map a section number -> the sheet that carries it (from the traversal).
    sheet_by_section: dict[int, object] = {}
    for idx, loc in enumerate(section_locations or []):
        name = _section_sheet_name(idx)
        if name not in wb.sheetnames:
            continue
        try:
            sheet_by_section.setdefault(int(loc.section), wb[name])
        except (TypeError, ValueError):
            continue

    def _put(ws, row: int, fp: dict) -> None:
        fnl, fsl = fp.get("fnl"), fp.get("fsl")
        fel, fwl = fp.get("fel"), fp.get("fwl")
        if fnl is not None:
            ws.cell(row, 9, round(fnl, 2))
            ws.cell(row, 10, 1)
        elif fsl is not None:
            ws.cell(row, 9, round(fsl, 2))
            ws.cell(row, 10, 2)
        if fel is not None:
            ws.cell(row, 11, round(fel, 2))
            ws.cell(row, 12, 1)
        elif fwl is not None:
            ws.cell(row, 11, round(fwl, 2))
            ws.cell(row, 12, 2)
        # Qtr-Qtr label (col M) — authoritative from the APD, else computed
        # from the survey position. Normalise "SE-SW"/"sesw" → "SESW".
        qq = fp.get("qq")
        if qq:
            ws.cell(row, 13, str(qq).replace("-", "").replace(" ", "").upper())

    for row, item in zip((8, 9, 10), dx_survey_footages):
        if not item:
            continue
        conc, fp = item
        if not fp or all(v is None for v in fp.values()):
            continue
        _put(shl, row, fp)  # summary sheet
        try:
            sec = PLSSKey.from_conc(conc).section
        except (ValueError, IndexError, AttributeError):
            sec = None
        own = sheet_by_section.get(sec)
        if own is not None and own is not shl:
            _put(own, row, fp)
    log.info("section_sheet.path_footages_written")


def _write_section_sheets_legacy(
    wb,
    design: CasingDesign,
    *,
    surface_location=None,
    producing_interval_location=None,
    td_location=None,
    intermediate_locations: list | None = None,
    plat_repo=None,
) -> None:
    """Original mapping used when no traversal is supplied.

    Layout: SHL = surface, BHL 1 = top of producing zone, BHL 3 = TD.
    ``intermediate_locations`` slots into BHL 2 (first item); additional
    items create BHL Section 4, 5, … by duplicating BHL Section 3.
    """
    intermediate_locations = list(intermediate_locations or [])
    bhl2_loc = intermediate_locations[0] if intermediate_locations else None
    extra_locs = intermediate_locations[1:]

    section_sheet_map = [
        ("SHL Section", surface_location),
        ("BHL Section 1", producing_interval_location),
        ("BHL Section 2", bhl2_loc),
        ("BHL Section 3", td_location),
    ]
    for sheet_name, location in section_sheet_map:
        if sheet_name in wb.sheetnames:
            _write_section_sheet(
                wb[sheet_name],
                design,
                location,
                sheet_label=sheet_name,
                plat_repo=plat_repo,
            )

    if extra_locs and "BHL Section 3" in wb.sheetnames:
        template_sheet = wb["BHL Section 3"]
        for i, loc in enumerate(extra_locs, start=4):
            new_name = f"BHL Section {i}"
            if new_name in wb.sheetnames:
                continue
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = new_name
            _write_section_sheet(
                new_sheet,
                design,
                loc,
                sheet_label=new_name,
                plat_repo=plat_repo,
            )
            log.info("section_sheet.created_dynamic", name=new_name)


def _gn_int(v) -> int | None:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _ensure_grid_numbers_coverage(gn_ws, section_locations, plat_repo) -> None:
    """Refresh the Grid Numbers sheet's geometry for every crossed section.

    Each section sheet draws its section outline — and places the well dot —
    from 16 "quarter-side" rows in this sheet (one per boundary segment),
    looked up by DGET on the section's PLSS key + side label. The embedded
    sheet ships a CURATED subset that is, in practice, badly incomplete:
    almost every section has only *some* of its 16 sides populated and the
    rest at zero length. A section with missing sides draws a collapsed
    outline, which is why the intermediate / producing-zone section sheets
    show no dot.

    So for every crossed section that has a plat polygon, we derive all 16
    quarter-sides from the polygon (``PlatRepository``) and write them in —
    OVERWRITING the section's existing rows in place (matched by PLSS key +
    side) and appending any side the sheet lacks. This makes each crossed
    section's outline complete and plat-exact. Sections with no plat polygon
    are left untouched (logged so a genuinely missing one is visible).
    """
    # Index existing rows by (6-int PLSS key, side label) so we can overwrite
    # the right row rather than appending duplicates the DGET would ignore.
    row_by_keyside: dict[tuple, int] = {}
    last_row = 2  # data starts at row 3 (rows 1-2 are headers)
    for r in range(3, gn_ws.max_row + 1):
        sec = gn_ws.cell(r, 1).value
        if sec is None:
            continue
        last_row = r
        key6 = tuple(_gn_int(gn_ws.cell(r, c).value) for c in range(1, 7))
        side = gn_ws.cell(r, 7).value
        row_by_keyside[(key6, side)] = r

    write_row = last_row + 1
    done: set[tuple] = set()
    for loc in section_locations:
        plss = PLSSKey.from_location(loc)
        if plss is None:
            continue
        key6 = (
            plss.section, plss.township, plss.township_dir,
            plss.range_, plss.range_dir, plss.baseline,
        )
        if key6 in done:
            continue
        done.add(key6)
        try:
            df = plat_repo._fetch_concs([plss.conc])  # noqa: SLF001
        except Exception as exc:
            log.warning("grid_numbers.fetch_failed", conc=plss.conc, error=str(exc))
            continue
        if df is None or df.empty:
            log.info("grid_numbers.no_plat_polygon", conc=plss.conc)
            continue
        pts = list(zip(df["Easting"].tolist(), df["Northing"].tolist()))
        rows = derive_section_corners(
            section=plss.section, township=plss.township,
            township_dir=plss.township_dir, range_=plss.range_,
            range_dir=plss.range_dir, baseline=plss.baseline, polygon_points=pts,
        )
        if not rows:
            continue
        refreshed = appended = 0
        for gc in rows:
            target = row_by_keyside.get((key6, gc.side))
            if target is None:
                target = write_row
                row_by_keyside[(key6, gc.side)] = target
                write_row += 1
                appended += 1
            else:
                refreshed += 1
            for col, val in enumerate(
                (
                    gc.section, gc.township, gc.township_dir, gc.range,
                    gc.range_dir, gc.baseline, gc.side, gc.length_ft,
                    gc.degrees, gc.minutes, gc.seconds, gc.alignment, gc.north_ref,
                ),
                start=1,
            ):
                gn_ws.cell(target, col, val)
        log.info(
            "grid_numbers.section_refreshed",
            conc=plss.conc, refreshed=refreshed, appended=appended,
        )


def _write_section_sheet(
    ws,
    design: CasingDesign,
    location,
    *,
    sheet_label: str,
    plat_repo=None,
) -> None:
    """Populate the well/API header, the PLSS input block at row 7, and
    the computed UTM coordinates (T7/U7/V7) for the location.

    ``plat_repo`` is an optional ``PlatRepository`` — when provided, we
    look up the section polygon and derive UTM from the APD footages
    via shapely geometry. With no plat polygon available the UTM cells
    stay at the template default.
    """
    ws["C2"] = design.well_name
    ws["C3"] = design.api
    if location is None:
        return

    # The SHL sheet and the BHL sheets read their PLSS direction inputs
    # DIFFERENTLY (confirmed against the reference workbook):
    #   * SHL   criteria are ``=$P$7``            → P7/R7 must be INT codes.
    #   * BHL   criteria are ``=IF($P$7="S",…)``  → P7/R7 must be the
    #                                               STRING "S"/"N", "W"/"E".
    # The original writer wrote int codes on every sheet, which made every
    # BHL DGET criterion fail (``IF(2="S",…)`` → wrong code) and the
    # bearing grid came up #VALUE!. Pick the encoding per sheet.
    is_shl = sheet_label.startswith("SHL")

    # Section / Township / Range / Meridian.
    if location.section:
        try:
            ws["N7"] = int(location.section)
        except ValueError:
            pass
    if location.township:
        try:
            ws["O7"] = int(location.township)
        except ValueError:
            pass
    if location.township_dir:
        d = location.township_dir.upper()
        ws["P7"] = (2 if d == "S" else 1) if is_shl else ("S" if d == "S" else "N")
    if location.range:
        try:
            ws["Q7"] = int(location.range)
        except ValueError:
            pass
    if location.range_dir:
        d = location.range_dir.upper()
        ws["R7"] = (2 if d == "W" else 1) if is_shl else ("W" if d == "W" else "E")
    if location.meridian:
        ws["S7"] = 2 if location.meridian.upper() == "U" else 1

    # On BHL sheets the section the bearing-grid DGET targets comes from
    # ``L38`` (an auto-detection formula), NOT N7. That detection only
    # sees the SHL + bottom-hole footages, so it can't resolve the
    # intermediate sections a lateral crosses. Force L38 to THIS section
    # so every crossed section's bearings resolve.
    if not is_shl and location.section:
        try:
            ws["L38"] = int(location.section)
        except ValueError:
            pass

    # Compute UTM from the APD's footages + the plat polygon. We also
    # write the four cardinal footages back into I7/K7 — those are
    # formula cells in the template referencing DxSurvey, but the user
    # wants the actual APD footages here.
    if plat_repo is None:
        return
    try:
        conc = _location_to_conc(location)
        if conc is None:
            return
        df = plat_repo._fetch_concs([conc])  # noqa: SLF001 — direct lookup
        if df.empty:
            log.info("section_sheet.plat_miss", sheet=sheet_label, conc=conc)
            return
        gdf = plat_repo._build_sections(df)  # noqa: SLF001
        if gdf.empty:
            return
        polygon = gdf.iloc[0].geometry
        fnl, fsl, fel, fwl = location_footages(location)
        if (fnl is None and fsl is None) or (fel is None and fwl is None):
            return
        x, y = footages_to_xy(polygon, fnl=fnl, fsl=fsl, fel=fel, fwl=fwl)
        ws["T7"] = round(x, 3)
        ws["U7"] = round(y, 3)
        ws["V7"] = 12  # UTM zone 12 for Utah

        # Also fill the four "Section Line Footages" cells so the user
        # sees the APD footages instead of the template's DxSurvey ref.
        # I7 = FNL or FSL (numeric); J7 = 1 if FNL, 2 if FSL
        # K7 = FEL or FWL (numeric); L7 = 1 if FEL, 2 if FWL
        if fnl is not None:
            ws["I7"] = fnl
            ws["J7"] = 1
        elif fsl is not None:
            ws["I7"] = fsl
            ws["J7"] = 2
        if fel is not None:
            ws["K7"] = fel
            ws["L7"] = 1
        elif fwl is not None:
            ws["K7"] = fwl
            ws["L7"] = 2
        log.info(
            "section_sheet.utm_written",
            sheet=sheet_label,
            conc=conc,
            utm=(round(x, 1), round(y, 1)),
        )
    except Exception as exc:
        log.warning(
            "section_sheet.utm_failed",
            sheet=sheet_label,
            location=location.name,
            error=str(exc),
        )


def _location_to_conc(location) -> str | None:
    """Build the 9-char Conc PLSS code (matches PlatRepository.BaseData).

    Format: ``"SSTTDRRRDDM"`` (2+2+1+2+1+1) → e.g. ``"2303S02WU"``.
    """
    try:
        sec = int(location.section)
        twp = int(location.township)
        rng = int(location.range)
    except (TypeError, ValueError):
        return None
    twpd = (location.township_dir or "").upper()
    rngd = (location.range_dir or "").upper()
    mer = (location.meridian or "").upper()
    if twpd not in ("N", "S") or rngd not in ("E", "W") or not mer:
        return None
    return f"{sec:02d}{twp:02d}{twpd}{rng:02d}{rngd}{mer}"


def _write_block(ws, top: int, s: CasingStringDesign, *, is_surface: bool) -> None:
    data_row = top + _DATA_ROW_OFFSET

    def put(col: str, row: int, value) -> None:
        if value is None:
            return
        ws[f"{col}{row}"] = value

    # Inputs
    put("B", data_row, s.hole_size_in)
    put("C", data_row, s.od_in)
    put("D", data_row, s.set_depth_md_ft)
    put("E", data_row, s.weight_ppf)
    put("F", data_row, s.grade)
    put("G", data_row, s.collar)
    put("H", data_row, s.cement_lead_sacks)
    put("I", data_row, s.cement_lead_yield)
    put("J", data_row, s.cement_tail_sacks)
    put("K", data_row, s.cement_tail_yield)

    # Engineering knobs
    put("B", top + 7, "y" if s.buoyed else "n")
    put("B", top + 8, s.mud_weight_ppg)
    # TVD (B19/B34/B49/B64). The template computes this via an exact-match
    # VLOOKUP of the set-depth MD into the DxSurvey survey table — but that
    # table still holds the template's *sample* survey (we populate the path
    # only at rows 8-10), so the lookup returns #N/A for every real well and
    # the error cascades into the per-string MASP/burst, the "Next" footer
    # rows (24/39/54/69) and the entire BOPE sheet. We already interpolate
    # the true TVD in Python, so write it directly and overwrite the formula.
    put("B", top + 9, s.set_depth_tvd_ft)
    put("B", top + 10, s.hole_washout_pct)
    put("B", top + 11, s.internal_gradient_psi_per_ft)
    put("B", top + 12, s.backup_mud_ppg)
    put("B", top + 13, s.internal_mud_ppg)

    # Computed values — Excel formulas will recompute these on open, but
    # we pre-fill so the workbook shows correct numbers even if formulas
    # haven't refreshed (e.g. headless openpyxl reads).
    put("Q", data_row, s.cement_height_ft)
    put("R", data_row, s.top_of_cement_ft)
    put("S", data_row, s.masp_psi)
    put("T", data_row, s.collapse_psi)
    put("U", data_row, s.collapse_load_psi)
    put("V", data_row, s.collapse_df)
    put("W", data_row, s.burst_psi)
    put("X", data_row, s.burst_load_psi)
    put("Y", data_row, s.burst_df)
    put("Z", data_row, s.joint_klbs)
    put("AA", data_row, s.tension_df)
    put("AB", data_row, s.neutral_point_ft)
    put("AC", data_row, s.tension_air_klbs)
    put("AD", data_row, s.tension_buoyed_klbs)
    put("AE", data_row, s.id_in)


def _write_footer_values(cr, design: CasingDesign) -> None:
    """Pre-fill each string's footer block (rows 24/39/54/69) with computed
    values, mirroring the template formulas, so they display even in viewers
    that don't recalc on open. Blank the footers of unused string slots.

    Per footer (for string ``i`` with next string ``n``), using the same
    0.05194806 psi/ft/ppg constant as the sheet:
        Next Set Depth (B)  = next string's setting TVD
        Next MW (D)         = next string's mud weight
        Next BHP (F)        = next MW * next TVD * 0.05194806
        Frac Init Press (I) = frac gradient * this string's TVD
        Max anticipated     = next BHP - (next TVD - this TVD) * next int.grad,
        pressure @ shoe (L)   or, for the last string, this TVD * MW * const.
    """
    K = 0.05194806
    footer_rows = (24, 39, 54, 69)
    frac = design.frac_gradient_psi_per_ft or 1.0
    n = len(design.strings)

    def put(row, col, val):
        cr.cell(row, col).value = None if val is None else round(val, 1)

    for idx in range(4):
        fr = footer_rows[idx]
        if idx >= n:  # unused slot — clear stale template constants
            for col in (2, 4, 6, 9, 12):
                cr.cell(fr, col).value = None
            continue
        cur = design.strings[idx]
        nxt = design.strings[idx + 1] if idx + 1 < n else None
        tvd = cur.set_depth_tvd_ft

        if tvd is not None:
            put(fr, 9, frac * tvd)  # I — fracture initiation pressure
            if nxt is not None and nxt.set_depth_tvd_ft is not None:
                n_bhp = nxt.mud_weight_ppg * nxt.set_depth_tvd_ft * K
                ig = nxt.internal_gradient_psi_per_ft or 0.0
                put(fr, 12, n_bhp - (nxt.set_depth_tvd_ft - tvd) * ig)  # L
            else:
                put(fr, 12, tvd * cur.mud_weight_ppg * K)  # L — BHP at shoe

        # Rows 24/39/54 carry Next Set Depth/MW/BHP; row 69 is a TOL row with
        # a different layout, so leave its B/D/F alone.
        if idx < 3:
            if nxt is not None and nxt.set_depth_tvd_ft is not None:
                put(fr, 2, nxt.set_depth_tvd_ft)                       # B
                cr.cell(fr, 4).value = nxt.mud_weight_ppg             # D
                put(fr, 6, nxt.mud_weight_ppg * nxt.set_depth_tvd_ft * K)  # F
            else:
                for col in (2, 4, 6):
                    cr.cell(fr, col).value = None


def _write_bope(ws, design: CasingDesign, *, bope_system_psi: float | None = None) -> None:
    """Fill the BOPE sheet's three hand-entered inputs.

    Everything else on the sheet is a formula off Casing Review (setting
    depth, MW, internal yield) that resolves once the per-string TVDs are
    written. The template ships these three as stale constants from whatever
    well it was built on; replace them with this well's values.

    Column layout: C=surface, D=intermediate, E=production, F=4th string.
    """
    from openpyxl.styles import Font

    review = build_bope_review(design, bope_system_psi=bope_system_psi)

    # Row 7 — Previous Shoe Setting Depth (TVD). D7/E7/F7 are formulas that
    # chain to the prior string's setting depth; only C7 (surface, no prior
    # casing) is a constant. Surface string has no previous shoe → 0.
    ws["C7"] = 0

    # Row 9 — BOPE Proposed (psi), per string. Permit-stated ratings are
    # written plain; inferred ratings are flagged in bold red.
    cols = ("C", "D", "E", "F")
    for idx, r in enumerate(review.strings[:4]):
        if r.bope_proposed_psi is not None:
            cell = ws[f"{cols[idx]}9"]
            cell.value = r.bope_proposed_psi
            if not r.bope_proposed_from_pdf:
                cell.font = Font(bold=True, color="FFCC0000")

    # Row 11 — Operator's Max Anticipated Pressure (psi). F11 converts it to
    # an equivalent mud weight via its own formula, so only C11 is an input.
    if review.operators_max_anticipated_pressure_psi is not None:
        ws["C11"] = round(review.operators_max_anticipated_pressure_psi, 1)

    # The F-column is the overflow slot for a 4th string. With no 4th string
    # its internal-yield ref ('Casing Review'!W57) is a DGET over empty inputs
    # = #VALUE!, which cascades into M15 and C50. Zero the input so the unused
    # slot stays clean instead of showing errors.
    if len(design.strings) < 4:
        ws["F10"] = 0


def _write_dataprint(ws, design: CasingDesign) -> None:
    """Write each string's normalized output into the DataPrint sheet.

    Column-range per string starts at column B (string 1), Q (string 2),
    AF (string 3), AU (string 4). The data rows start at row 11.
    """
    starts = ("B", "Q", "AF", "AU")
    for idx, s in enumerate(design.strings[:4]):
        col0 = starts[idx]
        # Row 7 carries the inch-prefix label, e.g. '9.625" Casing'.
        ws[f"{col0}7"] = f'{s.od_in}" Casing'
        # Row 11 starts the values; the spreadsheet repeats the per-stage
        # block but we just write the single shoe values for the lead row.
        row = 11
        _put(ws, col0, "C", row, s.masp_psi)
        _put(ws, col0, "D", row, s.collapse_psi)
        _put(ws, col0, "E", row, s.collapse_load_psi)
        _put(ws, col0, "F", row, s.collapse_df)
        _put(ws, col0, "G", row, s.burst_psi)
        _put(ws, col0, "H", row, s.burst_load_psi)
        _put(ws, col0, "I", row, s.burst_df)
        _put(ws, col0, "J", row, s.joint_klbs)
        _put(ws, col0, "K", row, s.tension_df)
        _put(ws, col0, "L", row, s.neutral_point_ft)
        _put(ws, col0, "M", row, s.tension_air_klbs)
        _put(ws, col0, "N", row, s.tension_buoyed_klbs)


def _put(ws, base_col: str, offset_col: str, row: int, value) -> None:
    """Write to (base_col + (offset_col - 'C') offset, row).

    The DataPrint panel uses C..N for the per-stage output columns within
    each string's block. We translate that to the string's starting column.
    """
    if value is None:
        return
    base_idx = openpyxl.utils.column_index_from_string(base_col)
    offset_idx = openpyxl.utils.column_index_from_string(offset_col)
    target = openpyxl.utils.get_column_letter(base_idx + offset_idx - 3)  # C is offset 0
    ws[f"{target}{row}"] = value
