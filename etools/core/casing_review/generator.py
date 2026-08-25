"""Fill the Casing Review xlsx template with values from an APD PDF.

The reference workbook has four STRING blocks of identical layout, each
starting at row ``10 + 15*(n-1)`` (so 10, 25, 40, 55). Within each block,
the operator-input cells live at fixed offsets from the block-top — that
lets us drive the whole sheet from a tiny ``CasingReviewInputs`` DTO.

Anything we don't fill is left to the spreadsheet formulas (TVD comes
from DxSurvey, MASP / Collapse / Burst / Tension all recompute via DGET
lookups against the Casing Strengths sheet).
"""

from __future__ import annotations

from etools.core.io_safety import atomic_output
import shutil
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

CASING_REVIEW_TEMPLATE = Path(__file__).parent / "templates" / "casing_review_template.xlsx"

# Row offsets relative to the STRING-block top (row 10 for STRING 1).
# Cells without an offset are formulas and stay untouched.
_INPUT_CELL = "B"  # most input cells live in column B
_HEADER_INPUT_ROW = 2  # +2 → row 12 for STRING 1; the "row with the data"

# Default safety-related inputs the APD doesn't carry — Newfield's
# engineering-review convention. Surface string uses higher hole washout
# and lower internal gradient than deeper strings.
_DEFAULTS_SURFACE = {
    "hole_washout_pct": 10.0,
    "internal_gradient_psi_per_ft": 0.12,
    "backup_mud_ppg": 0.0,
    "internal_mud_ppg": 0.0,
    "buoyed": "y",
}
_DEFAULTS_DEEP = {
    "hole_washout_pct": 4.0,
    "internal_gradient_psi_per_ft": 0.22,
    "backup_mud_ppg": 0.0,
    "internal_mud_ppg": 0.0,
    "buoyed": "y",
}


@dataclass
class StringInputs:
    """One casing-string row from the APD's Hole/Casing/Cement table."""

    hole_size_in: float | None = None
    casing_size_in: float | None = None
    set_depth_ft: float | None = None  # bottom of the string
    weight_ppf: float | None = None
    grade: str | None = None  # e.g. "J-55"
    collar: str | None = None  # "BTC" / "STC" / "LTC" / etc.
    # Cement — typically two stages: lead (Type II / Class G) + tail
    # (Premium Plus). APD usually supplies sacks and yield for each.
    cement_lead_sacks: int | None = None
    cement_lead_yield: float | None = None
    cement_tail_sacks: int | None = None
    cement_tail_yield: float | None = None
    mud_weight_ppg: float | None = None
    # Engineering knobs the APD doesn't carry. ``None`` ⇒ use the
    # surface-vs-deep default for this string.
    hole_washout_pct: float | None = None
    internal_gradient_psi_per_ft: float | None = None
    backup_mud_ppg: float | None = None
    internal_mud_ppg: float | None = None
    buoyed: str | None = None  # "y" / "n"


@dataclass
class CasingReviewInputs:
    """Everything we feed the Casing Review template."""

    company: str | None = None
    well_name: str | None = None
    api: str | None = None
    frac_gradient_psi_per_ft: float = 1.0
    # Up to four strings in order: surface, intermediate, production, liner.
    # Surface is index 0; downhole strings get the "deep" defaults.
    strings: list[StringInputs] = field(default_factory=list)


# Block-top rows for STRING 1..4 in the reference template.
_STRING_BLOCK_TOP_ROWS = (10, 25, 40, 55)


def generate_casing_review(
    inputs: CasingReviewInputs,
    output_path: Path,
    *,
    template_path: Path | None = None,
) -> Path:
    """Fill the template with ``inputs`` and save to ``output_path``, atomically.

    Built at a sibling temp path and swapped into place only on success, so a
    failure part-way through cannot destroy a previously generated workbook.
    """
    output_path = Path(output_path)
    with atomic_output(output_path) as work_path:
        _generate_casing_review_to(
            inputs, work_path, template_path=template_path
        )
    return output_path


def _generate_casing_review_to(
    inputs: CasingReviewInputs,
    output_path: Path,
    *,
    template_path: Path | None = None,
) -> Path:
    """Fill the template at ``template_path`` with ``inputs`` and save to
    ``output_path``. Returns the resolved output path."""
    template_path = Path(template_path or CASING_REVIEW_TEMPLATE)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy the template byte-for-byte first, then open and overwrite the
    # input cells. This preserves charts, conditional formatting, named
    # ranges, and the Casing Strengths lookup that openpyxl can't always
    # round-trip cleanly.
    shutil.copyfile(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Casing Review"]

    if inputs.company is not None:
        ws["B4"] = inputs.company
    if inputs.well_name is not None:
        ws["B5"] = inputs.well_name
    if inputs.api is not None:
        ws["B6"] = inputs.api
    if inputs.frac_gradient_psi_per_ft is not None:
        ws["B9"] = inputs.frac_gradient_psi_per_ft

    for idx, s in enumerate(inputs.strings[:4]):
        top = _STRING_BLOCK_TOP_ROWS[idx]
        defaults = _DEFAULTS_SURFACE if idx == 0 else _DEFAULTS_DEEP
        _write_string_block(ws, top, s, defaults)

    wb.save(output_path)
    return output_path


def _write_string_block(ws, top: int, s: StringInputs, defaults: dict) -> None:
    """Write a single STRING block. ``top`` is the block header row
    (10 / 25 / 40 / 55). The data row is ``top + 2``; engineering knobs
    live at ``top + 7`` through ``top + 13``."""
    data_row = top + 2

    def put(col: str, row: int, value) -> None:
        if value is None:
            return
        ws[f"{col}{row}"] = value

    put("B", data_row, s.hole_size_in)
    put("C", data_row, s.casing_size_in)
    put("D", data_row, s.set_depth_ft)
    put("E", data_row, s.weight_ppf)
    put("F", data_row, s.grade)
    put("G", data_row, s.collar)
    put("H", data_row, s.cement_lead_sacks)
    put("I", data_row, s.cement_lead_yield)
    put("J", data_row, s.cement_tail_sacks)
    put("K", data_row, s.cement_tail_yield)

    # Engineering knobs (rows 17-23 for STRING 1 ⇒ top+7..top+13).
    put("B", top + 7, s.buoyed if s.buoyed is not None else defaults["buoyed"])
    put("B", top + 8, s.mud_weight_ppg)
    put(
        "B",
        top + 10,
        s.hole_washout_pct
        if s.hole_washout_pct is not None
        else defaults["hole_washout_pct"],
    )
    put(
        "B",
        top + 11,
        s.internal_gradient_psi_per_ft
        if s.internal_gradient_psi_per_ft is not None
        else defaults["internal_gradient_psi_per_ft"],
    )
    put(
        "B",
        top + 12,
        s.backup_mud_ppg
        if s.backup_mud_ppg is not None
        else defaults["backup_mud_ppg"],
    )
    put(
        "B",
        top + 13,
        s.internal_mud_ppg
        if s.internal_mud_ppg is not None
        else defaults["internal_mud_ppg"],
    )
