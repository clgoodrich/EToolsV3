"""WCR Excel writer.

Output schema (matches ``tests/South_Moon_5-31-32-C4-3H_4301353996_WCR.xlsx``):

    Row 1 :  WellName | <value>
    Row 2 :  API | <value>
    Row 3 :  Operator | <value>
    Row 4 :  WellType | <value>
    Row 5 :  SpudDate | <value>
    Row 6 :  RotaryRigDate | <value>
    Row 7 :  TDReachedDate | <value>
    Row 8 :  CompletedOrAbandonedDate | <value>
    Row 9 :  (header) MeasuredDepth | TVD | Easting | Northing | FNL | FSL |
             FEL | FWL | Section | Township | Township_Direction | Range |
             Range_Direction | Baseline
    Rows 10-14 : SHL | Control_Point | Frac_Start | Frac_End | BHL

Optional blocks (match the hand-made V2-era workbooks, e.g. the Reay WCR):

    E1:G2  : Perf Top | Perf Bottom | Perf Date header + values
    Row 16 : casing header — Feature | Top | Bottom | Diam | Weight | Grade |
             Connection Type | Cement Top | Cement Bottom | Cement Type |
             Sacks | Yield | Cement Weight
    Row 17+: one row per hole/casing/cement record

All numbers are written as numbers (not strings).
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook

from etools.core.io_safety import atomic_output
from etools.logging_setup import get_logger
from etools.models import WCRLocationRow, WCRWellInfo

log = get_logger(__name__)


_INFO_LABELS: tuple[tuple[str, str], ...] = (
    ("WellName", "well_name"),
    ("API", "api_well_no"),
    ("Operator", "operator"),
    ("WellType", "well_type"),
    ("SpudDate", "spud_date"),
    ("RotaryRigDate", "rotary_date"),
    ("TDReachedDate", "td_date"),
    ("CompletedOrAbandonedDate", "completion_date"),
)

_FOOTAGE_HEADERS: tuple[str, ...] = (
    "",
    "MeasuredDepth",
    "TVD",
    "Easting",
    "Northing",
    "FNL",
    "FSL",
    "FEL",
    "FWL",
    "Section",
    "Township",
    "Township_Direction",
    "Range",
    "Range_Direction",
    "Baseline",
)

_LOCATION_ORDER: tuple[str, ...] = (
    "SHL",
    "Control_Point",
    "Frac_Start",
    "Frac_End",
    "BHL",
)

_CASING_HEADERS: tuple[str, ...] = (
    "Feature",
    "Top",
    "Bottom",
    "Diam",
    "Weight",
    "Grade",
    "Connection Type",
    "Cement Top",
    "Cement Bottom",
    "Cement Type",
    "Sacks",
    "Yield",
    "Cement Weight",
)

# Repository column -> output column position (matches _CASING_HEADERS order).
_CASING_SOURCE_COLS: tuple[str, ...] = (
    "Feature",
    "Top_MD",
    "Bottom_MD",
    "Diameter",
    "Weight",
    "Grade",
    "ConnectionType",
    "CementTop",
    "CementBottom",
    "CementType",
    "Sacks",
    "Yield",
    "CementWeight",
)

_CASING_HEADER_ROW = 16


def generate_wcr_excel(
    *,
    info: WCRWellInfo,
    location_rows: Iterable[WCRLocationRow],
    output_path: str | Path,
    perf_top_md: float | None = None,
    perf_bottom_md: float | None = None,
    perf_date: str | None = None,
    casing: pd.DataFrame | None = None,
) -> Path:
    """Write the WCR workbook to ``output_path``, atomically.

    The workbook is built in memory and saved to a sibling temp path, which is
    swapped into place only on success -- so a failed save (most often the file
    being open in Excel) leaves the previous workbook intact rather than
    truncating it.
    """
    out_path = Path(output_path)
    with atomic_output(out_path) as work_path:
        _generate_wcr_excel_to(
            info=info,
            location_rows=location_rows,
            output_path=work_path,
            perf_top_md=perf_top_md,
            perf_bottom_md=perf_bottom_md,
            perf_date=perf_date,
            casing=casing,
        )
    return out_path


def _generate_wcr_excel_to(
    *,
    info: WCRWellInfo,
    location_rows: Iterable[WCRLocationRow],
    output_path: str | Path,
    perf_top_md: float | None = None,
    perf_bottom_md: float | None = None,
    perf_date: str | None = None,
    casing: pd.DataFrame | None = None,
) -> Path:
    """Write the WCR workbook to ``output_path``. Overwrites if it exists.

    ``casing`` takes the WCRRepository frame shape (Feature/Top_MD/.../CementWeight).
    Perf values and casing rows are optional — omitted blocks leave those
    cells blank, matching the South Moon reference layout.
    """
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"

    _write_info(ws, info)
    if perf_top_md is not None or perf_bottom_md is not None or perf_date:
        _write_perf_block(ws, perf_top_md, perf_bottom_md, perf_date)
    _write_headers(ws)

    by_name = {row.name: row for row in location_rows}
    for offset, name in enumerate(_LOCATION_ORDER):
        row = by_name.get(name)
        r = 10 + offset
        ws.cell(row=r, column=1, value=name)
        if row is None:
            continue
        _write_location(ws, row, r)

    if casing is not None and len(casing):
        _write_casing(ws, casing)

    wb.save(out_path)
    log.info("wcr.excel.saved", path=str(out_path))
    return out_path


# ---------------------------------------------------------------------------
# Section writers
# ---------------------------------------------------------------------------


def _write_info(ws, info: WCRWellInfo) -> None:
    api = (info.api_well_no or "")[:10]
    for r, (label, attr) in enumerate(_INFO_LABELS, start=1):
        ws.cell(row=r, column=1, value=label)
        value = api if attr == "api_well_no" else getattr(info, attr, None)
        ws.cell(row=r, column=2, value=_serialize(value))


def _write_headers(ws) -> None:
    for col, label in enumerate(_FOOTAGE_HEADERS, start=1):
        if label:  # leave column A header blank to match the reference output
            ws.cell(row=9, column=col, value=label)


def _write_perf_block(
    ws, top_md: float | None, bottom_md: float | None, perf_date: str | None
) -> None:
    ws.cell(row=1, column=5, value="Perf Top")
    ws.cell(row=1, column=6, value="Perf Bottom")
    ws.cell(row=1, column=7, value="Perf Date")
    ws.cell(row=2, column=5, value=_round(top_md, 0))
    ws.cell(row=2, column=6, value=_round(bottom_md, 0))
    ws.cell(row=2, column=7, value=perf_date)


def _write_casing(ws, casing: pd.DataFrame) -> None:
    for col, label in enumerate(_CASING_HEADERS, start=1):
        ws.cell(row=_CASING_HEADER_ROW, column=col, value=label)
    for i, (_, src) in enumerate(casing.iterrows()):
        r = _CASING_HEADER_ROW + 1 + i
        for col, name in enumerate(_CASING_SOURCE_COLS, start=1):
            value = src.get(name)
            if value is None or pd.isna(value):
                continue
            if isinstance(value, (int, float)):
                value = round(float(value), 2)
                if value == int(value):
                    value = int(value)
            ws.cell(row=r, column=col, value=_serialize(value))


def _write_location(ws, row: WCRLocationRow, r: int) -> None:
    ws.cell(row=r, column=2, value=_round(row.measured_depth, 0))
    ws.cell(row=r, column=3, value=_round(row.tvd, 2))
    ws.cell(row=r, column=4, value=_round(row.easting, 0))
    ws.cell(row=r, column=5, value=_round(row.northing, 0))
    ws.cell(row=r, column=6, value=_round(row.fnl, 2))
    ws.cell(row=r, column=7, value=_round(row.fsl, 2))
    ws.cell(row=r, column=8, value=_round(row.fel, 2))
    ws.cell(row=r, column=9, value=_round(row.fwl, 2))
    ws.cell(row=r, column=10, value=row.section)
    ws.cell(row=r, column=11, value=row.township)
    ws.cell(row=r, column=12, value=row.township_dir)
    ws.cell(row=r, column=13, value=row.range)
    ws.cell(row=r, column=14, value=row.range_dir)
    ws.cell(row=r, column=15, value=row.baseline)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _serialize(value):
    """Render dates as ``YYYY-MM-DD`` to match the reference output, pass everything else through."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def _round(value, ndigits: int):
    if value is None:
        return None
    try:
        rounded = round(float(value), ndigits)
        # Excel prefers integer when ndigits == 0
        return int(rounded) if ndigits == 0 else rounded
    except (TypeError, ValueError):
        return None
